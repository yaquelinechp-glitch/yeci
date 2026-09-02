import os
import uuid
import json
import base64
import hashlib
import hmac
import random
from datetime import datetime, timezone, timedelta

import bcrypt
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.utils import timezone as dj_timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken
from django.db.models import Sum, Q, Count, F, Avg

from .models import (
    Partner, Course, CourseVideo, QuizQuestion, QuizBankQuestion, CourseAssignment,
    CourseRating, QuizAttempt, PartnerProgress,
    Deal, Commission, Opportunity, OpportunityEvent, TrainingResult, Certification,
    TokenBlacklist, LoginAttempt, PartnerOnboarding, PartnerUser, MdfRequest,
    Reward, RewardRedemption, PointTransaction, ChannelConflict,
    Communication, CommunicationRecipient, Product, Notification,
    CourseExamQuestion, CostExportSetting,
)
from . import security
from .rewards import (
    award_points, spend_points, refund_points, adjust_points, point_serialize,
)
from .onboarding import onboarding_snapshot, STEP_ORDER
from .serializers import (
    PartnerSerializer, PartnerCreateSerializer, PartnerUpdateSerializer,
    CourseSerializer, CourseCreateSerializer, CourseUpdateSerializer, CourseVideoSerializer,
    QuizQuestionSerializer, QuizQuestionAdminSerializer, QuizAttemptSerializer,
    QuizBankQuestionSerializer, CourseAssignmentSerializer, CourseRatingSerializer,
    DealSerializer, DealCreateSerializer, DealUpdateSerializer,
    CommissionSerializer,
    OpportunitySerializer, OpportunityCreateSerializer, OpportunityUpdateSerializer,
    OpportunityEventSerializer, TrainingResultSerializer, CertificationSerializer,
    ProductSerializer, NotificationSerializer, CourseExamQuestionSerializer,
)


# ─── Auth helpers ───────────────────────────────────────

EURO_CURRENCY_COUNTRIES = [
    "at", "be", "bg", "hr", "cy", "cz", "dk", "ee", "fi", "fr", "de", "gr",
    "hu", "ie", "it", "lv", "lt", "lu", "mt", "nl", "pl", "pt", "ro", "sk",
    "si", "es", "se", "gb", "no", "is", "li",
]

def currency_for_country(country: str = "") -> str:
    """Derive currency from the partner's country. Europe -> EUR, Switzerland -> CHF, default USD."""
    code = (country or "").strip().lower()
    if code in ("ch", "switzerland", "suiza"):
        return "chf"
    if code in EURO_CURRENCY_COUNTRIES:
        return "eur"
    return "usd"


def _make_user_response(user, member=None):
    """Build JWT + user payload for a Partner (optionally acting as a PartnerUser member)."""
    refresh = RefreshToken()
    refresh["user_id"] = user.id
    refresh["role"] = user.role
    if member:
        refresh["member_id"] = member.id
    access_token = str(refresh.access_token)
    refresh_token = str(refresh)
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "user": {
            "id": user.id, "company_name": user.company_name, "email": (member.email if member else user.email),
            "phone": user.phone, "tax_id": user.tax_id,
            "country": user.country or "",
            "contact_name": (member.contact_name if member else user.contact_name),
            "first_name": (member.first_name if member else user.first_name) or "",
            "last_name": (member.last_name if member else user.last_name) or "",
            "username": (member.username if member else user.username) or "",
            "avatar": (member.avatar if member else user.avatar) or "",
            "role": user.role, "status": user.status, "training_track": user.training_track,
            "commission_rate": user.commission_rate,
            "certification_date": user.certification_date.isoformat() if user.certification_date else None,
            "why_partner": user.why_partner or "", "sales_approach": user.sales_approach or "",
            "created_at": user.created_at.isoformat() if user.created_at else "",
            "member": bool(member),
            "member_id": (member.id if member else None),
            "member_role": (member.role if member else None),
        },
    }


def _verify_google_id_token(id_token: str) -> dict | None:
    """
    Decode and verify a Google ID token.
    In production, verify the signature with Google's public keys.
    For this demo, we decode and validate the basic claims.
    """
    try:
        parts = id_token.split(".")
        if len(parts) != 3:
            return None
        payload_b64 = parts[1]
        # Fix base64 padding
        padding = 4 - len(payload_b64) % 4
        if padding != 4:
            payload_b64 += "=" * padding
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        # Validate basic claims
        if not payload.get("email"):
            return None
        # Check expiration
        import time
        if payload.get("exp", 0) < time.time():
            return None
        # Check issuer
        if payload.get("iss") not in ("accounts.google.com", "https://accounts.google.com"):
            return None
        return payload
    except Exception:
        return None


# ─── Auth views ─────────────────────────────────────────

@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    email = request.data.get("email", "")
    password = request.data.get("password", "")
    if not email or not password:
        return Response({"detail": "Email and password are required"}, status=400)

    ip = request.META.get("REMOTE_ADDR", "0.0.0.0")
    if security.is_rate_limited(ip):
        return Response({"detail": "Demasiados intentos. Intenta de nuevo en 15 minutos."}, status=429)

    try:
        user = Partner.objects.get(email=email)
    except Partner.DoesNotExist:
        member = PartnerUser.objects.filter(email=email).first()
        if not member:
            security.record_login_attempt(ip, False)
            return Response({"detail": "Credenciales incorrectas"}, status=401)
        if member.status != "activo":
            security.record_login_attempt(ip, False)
            return Response({"detail": "Cuenta no activa. Acepta la invitación o contacta con tu partner."}, status=403)
        if not member.password_hash or not bcrypt.checkpw(password.encode(), member.password_hash.encode()):
            security.record_login_attempt(ip, False)
            return Response({"detail": "Credenciales incorrectas"}, status=401)
        partner = member.partner
        if partner.status != "activo":
            security.record_login_attempt(ip, False)
            return Response({"detail": "El partner no está activo"}, status=403)
    security.record_login_attempt(ip, True)
    return Response(_make_user_response(user))


@api_view(["GET", "PUT"])
def profile(request):
    """Get or update the current user's profile (name, username, avatar)."""
    user = _current_user(request)
    if not user:
        return Response({"detail": "Authentication required"}, status=401)

    target = user.member if getattr(user, "is_member", False) else user

    if request.method == "GET":
        return Response({
            "first_name": target.first_name or "",
            "last_name": target.last_name or "",
            "username": target.username or "",
            "avatar": target.avatar or "",
            "contact_name": target.contact_name or "",
            "company_name": user.company_name or "",
            "email": user.email or "",
            "country": user.country or "",
        })

    data = request.data or {}
    for f in ("first_name", "last_name", "username"):
        if f in data:
            setattr(target, f, (str(data.get(f) or "")).strip()[:200])
    if "country" in data:
        user.country = ((data.get("country") or "").strip())[:100]
    user.save(update_fields=["first_name", "last_name", "username", "country"])
    if "avatar" in data:
        av = (data.get("avatar") or "").strip()
        if av and not av.startswith("data:image/"):
            return Response({"detail": "Avatar must be a data:image URL"}, status=400)
        if len(av) > 2_500_000:
            return Response({"detail": "Avatar image is too large"}, status=400)
        target.avatar = av
    target.save(update_fields=["first_name", "last_name", "username", "avatar"])
    return Response({"ok": True})


@api_view(["POST"])
@permission_classes([AllowAny])
def register_view(request):
    ser = PartnerCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data

    if Partner.objects.filter(email=d["email"]).exists() or PartnerUser.objects.filter(email=d["email"]).exists():
        return Response({"detail": "El email ya esta registrado"}, status=400)

    password = d.get("password", "")
    password_err = security.validate_password_strength(password)
    if password_err:
        return Response({"detail": password_err}, status=400)
    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

    partner = Partner(
        company_name=d["company_name"], email=d["email"],
        password_hash=password_hash,
        phone=d.get("phone", ""), tax_id=d.get("tax_id", ""),
        country=d.get("country", ""),
        contact_name=d.get("contact_name", ""),
        why_partner=d.get("why_partner", ""), sales_approach=d.get("sales_approach", ""),
    )
    partner.save()
    return Response(_make_user_response(partner), status=201)


@api_view(["POST"])
@permission_classes([AllowAny])
def logout_view(request):
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.split(" ", 1)[1]
        try:
            from rest_framework_simplejwt.tokens import AccessToken
            at = AccessToken(token)
            TokenBlacklist.objects.get_or_create(jti=at["jti"], user_id=at.get("user_id", ""))
        except Exception:
            pass
    return Response({"detail": "Sesión cerrada"}, status=200)


@api_view(["POST"])
@permission_classes([AllowAny])
def google_auth_view(request):
    """
    Authenticate or register a user via Google.
    Accepts: { credential: "<Google ID token>" }
    Returns: { access_token, refresh_token, user }
    """
    credential = request.data.get("credential", "")
    if not credential:
        return Response({"detail": "Google credential is required"}, status=400)

    payload = _verify_google_id_token(credential)
    if not payload:
        return Response({"detail": "Invalid Google credential"}, status=400)

    google_email = payload.get("email", "").lower().strip()
    google_id = payload.get("sub", "")
    google_name = payload.get("name", "")
    google_picture = payload.get("picture", "")

    if not google_email or not google_id:
        return Response({"detail": "Invalid Google credential: missing email or sub"}, status=400)

    # Find or create user
    try:
        user = Partner.objects.get(email=google_email)
        # User exists — link Google account if not already linked
        if user.provider != "google":
            user.provider = "google"
            user.google_id = google_id
            user.save(update_fields=["provider", "google_id"])
    except Partner.DoesNotExist:
        # Create new partner from Google data
        company_name = google_name or google_email.split("@")[0]
        user = Partner(
            company_name=company_name,
            email=google_email,
            contact_name=google_name,
            provider="google",
            google_id=google_id,
        )
        user.save()

    return Response(_make_user_response(user))


# ─── Partners ───────────────────────────────────────────

@api_view(["GET"])
def list_partners(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    partners = Partner.objects.exclude(role="admin").order_by("-created_at")
    return Response(PartnerSerializer(partners, many=True).data)


@api_view(["GET"])
def list_solicitudes(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    partners = Partner.objects.exclude(role="admin").filter(
        status__in=["solicitado", "en_revision"]
    ).order_by("-created_at")
    return Response(PartnerSerializer(partners, many=True).data)


def _ensure_compliance_assignment(partner):
    """RQ-14: on partner activation, auto-assign the Compliance Track with a 30-day deadline."""
    course = Course.objects.filter(track="cumplimiento", status__in=["publicado", ""]).first()
    if not course:
        return None
    deadline = dj_timezone.localdate() + timedelta(days=30)
    asg, _ = CourseAssignment.objects.get_or_create(
        course=course, partner=partner,
        defaults={"deadline": deadline},
    )
    if asg.deadline is None or asg.deadline > deadline:
        asg.deadline = deadline
        asg.save(update_fields=["deadline"])
    return asg


@api_view(["GET", "PATCH"])
def partner_detail(request, partner_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.id != partner_id:
        return Response({"detail": "Access denied"}, status=403)
    if request.method != "GET" and getattr(user, "is_member", False):
        return Response({"detail": "Solo el owner del partner puede editar el perfil"}, status=403)

    try:
        partner = Partner.objects.get(id=partner_id)
    except Partner.DoesNotExist:
        return Response({"detail": "Partner not found"}, status=404)

    if request.method == "GET":
        return Response(PartnerSerializer(partner).data)

    ser = PartnerUpdateSerializer(data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data
    for k, v in d.items():
        setattr(partner, k, v)
    partner.save()
    if partner.status == "activo":
        _ensure_compliance_assignment(partner)
        _notify(
            [partner.id], "cuenta",
            {"en": "Welcome to the Partner Portal", "es": "Bienvenido al Portal de Partners", "de": "Willkommen im Partner-Portal"},
            {"en": "Your account has been activated. Start exploring.", "es": "Tu cuenta ha sido activada. Empieza a explorar.", "de": "Ihr Konto wurde aktiviert. Legen Sie los."},
            "/partner",
        )
    return Response(PartnerSerializer(partner).data)


# ─── Courses ────────────────────────────────────────────

def _visible_courses(user):
    """Role-based course visibility: every active partner sees all published courses plus any explicitly assigned."""
    if user.role == "admin":
        return Course.objects.all()
    published = Course.objects.filter(status__in=["publicado", ""])
    assigned = Course.objects.filter(assignments__partner=user)
    combined = {}
    for c in list(published) + list(assigned):
        combined[c.id] = c
    return list(combined.values())


@api_view(["GET", "POST"])
def courses_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    lang = request.query_params.get("lang", "en")

    if request.method == "GET":
        if user.role == "socio" and user.status != "activo":
            return Response({"detail": "Solo partners aprobados pueden ver cursos"}, status=403)
        partner = user if user.role == "socio" else None
        courses = _visible_courses(user)
        return Response(CourseSerializer(courses, many=True, partner=partner, lang=lang).data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    ser = CourseCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    course = Course(**ser.validated_data)
    course.save()
    title = course.title or {}
    if not isinstance(title, dict):
        title = {"en": str(title), "es": str(title), "de": str(title)}
    _notify_all_active(
        "curso_nuevo",
        title,
        {"en": "A new course is available in the portal", "es": "Hay un nuevo curso disponible en el portal", "de": "Ein neuer Kurs ist im Portal verfügbar"},
        "/partner/courses",
    )
    return Response({"id": course.id, "title": course.title})


@api_view(["GET"])
def list_courses(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados pueden ver cursos"}, status=403)

    lang = request.query_params.get("lang", "en")
    partner = user if user.role == "socio" else None
    courses = _visible_courses(user)
    data = CourseSerializer(courses, many=True, partner=partner, lang=lang).data
    return Response(data)


@api_view(["GET", "PATCH", "DELETE"])
def course_detail(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    lang = request.query_params.get("lang", "en")

    if request.method == "GET":
        partner = user if user.role == "socio" else None
        return Response(CourseSerializer(course, partner=partner, lang=lang).data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    if request.method == "PATCH":
        ser = CourseUpdateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        for k, v in ser.validated_data.items():
            setattr(course, k, v)
        course.save()
        return Response(CourseSerializer(course, lang=lang).data)

    if request.method == "DELETE":
        for v in course.videos.all():
            rel = v.video_url
            if rel.startswith(settings.MEDIA_URL):
                rel = rel[len(settings.MEDIA_URL):]
            fp = os.path.join(settings.MEDIA_ROOT, rel)
            if os.path.exists(fp):
                os.remove(fp)
        course.delete()
        return Response({"ok": True})


@api_view(["POST"])
def create_course(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    ser = CourseCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    course = Course(**ser.validated_data)
    course.save()
    return Response({"id": course.id, "title": course.title})


# ─── Course Videos ──────────────────────────────────────

@api_view(["POST"])
def upload_video(request, course_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    file = request.FILES.get("file")
    url_param = (request.query_params.get("video_url", "") or "").strip()
    if not file and not url_param:
        return Response({"detail": "No file or url provided"}, status=400)

    video_url = ""
    if file:
        upload_dir = os.path.join(settings.MEDIA_ROOT, "videos")
        os.makedirs(upload_dir, exist_ok=True)

        ext = os.path.splitext(file.name)[1] if file.name else ".mp4"
        filename = f"{uuid.uuid4().hex}{ext}"
        filepath = os.path.join(upload_dir, filename)
        with open(filepath, "wb") as f:
            for chunk in file.chunks():
                f.write(chunk)
        video_url = f"/uploads/videos/{filename}"
    else:
        video_url = url_param

    title_raw = request.query_params.get("title", "") or (file.name if file else "") or "Video"
    description_raw = request.query_params.get("description", "")
    video_order = int(request.query_params.get("video_order", -1))
    if video_order < 0:
        video_order = CourseVideo.objects.filter(course=course).count()
    phase = int(request.query_params.get("phase", 1))
    day = int(request.query_params.get("day", 1))

    title = {"en": title_raw, "es": title_raw, "de": title_raw}
    description = {"en": description_raw, "es": description_raw, "de": description_raw}

    video = CourseVideo(
        course=course, title=title, description=description,
        video_url=video_url, video_order=video_order,
        phase=phase, day=day,
    )
    video.save()
    return Response({"id": video.id, "title": video.title, "video_url": video.video_url, "video_order": video.video_order, "phase": video.phase, "day": video.day})


@api_view(["POST"])
def upload_material(request, course_id=None):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    if course_id:
        try:
            Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return Response({"detail": "Course not found"}, status=404)

    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file provided"}, status=400)

    upload_dir = os.path.join(settings.MEDIA_ROOT, "materials")
    os.makedirs(upload_dir, exist_ok=True)

    ext = os.path.splitext(file.name)[1] if file.name else ".bin"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)

    name_raw = (request.query_params.get("name") or "").strip() or file.name or "Material"
    material = {"id": uuid.uuid4().hex[:8], "name": name_raw, "url": f"/uploads/materials/{filename}"}
    return Response(material)


@api_view(["POST"])
def upload_thumbnail(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file provided"}, status=400)

    upload_dir = os.path.join(settings.MEDIA_ROOT, "thumbnails")
    os.makedirs(upload_dir, exist_ok=True)

    ext = os.path.splitext(file.name)[1] if file.name else ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)

    return Response({"url": f"/uploads/thumbnails/{filename}"})


@api_view(["DELETE"])
def delete_video(request, course_id, video_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
    except CourseVideo.DoesNotExist:
        return Response({"detail": "Video not found"}, status=404)

    rel = video.video_url
    if rel.startswith(settings.MEDIA_URL):
        rel = rel[len(settings.MEDIA_URL):]
    fp = os.path.join(settings.MEDIA_ROOT, rel)
    if os.path.exists(fp):
        os.remove(fp)
    video.delete()
    return Response({"ok": True})


@api_view(["POST"])
def reorder_videos(request, course_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    order = request.data.get("order", [])
    if not order:
        return Response({"detail": "No order provided"}, status=400)
    for idx, vid in enumerate(order):
        CourseVideo.objects.filter(id=vid, course_id=course_id).update(video_order=idx)
    return Response({"ok": True})


@api_view(["POST"])
def duplicate_course(request, course_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        src = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    title = dict(src.title or {})
    if isinstance(title, dict):
        for k in list(title.keys()):
            title[k] = f"{title[k]} (copy)" if title[k] else title[k]

    copy = Course(
        title=title,
        description=dict(src.description or {}),
        thumbnail_url=src.thumbnail_url,
        category=src.category, level=src.level, track=src.track,
        status="borrador",
        related_products=list(src.related_products or []),
        pass_mark=src.pass_mark, validity_months=src.validity_months,
        prerequisite_course_id="",
        quiz_questions_count=src.quiz_questions_count,
        exam_questions_count=src.exam_questions_count,
        materials=list(src.materials or []),
        phase_config=list(src.phase_config or []),
    )
    copy.save()

    for v in src.videos.all():
        nv = CourseVideo(
            course=copy,
            title=dict(v.title or {}), description=dict(v.description or {}),
            video_url=v.video_url, duration_seconds=v.duration_seconds,
            video_order=v.video_order, phase=v.phase, day=v.day,
        )
        nv.save()
        for q in v.quiz_questions.all():
            QuizQuestion.objects.create(
                video=nv,
                question=dict(q.question or {}),
                options=list(q.options or []),
                correct_index=q.correct_index,
                correct_indices=list(q.correct_indices or []),
                fill_answer=dict(q.fill_answer or {}),
                question_type=q.question_type or "single",
                order=q.order,
            )
    return Response({"id": copy.id, "title": copy.title})


# ─── Question Bank ──────────────────────────────────────

@api_view(["GET", "POST"])
def quiz_bank(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        questions = QuizBankQuestion.objects.all()
        data = QuizBankQuestionSerializer(questions, many=True).data
        return Response(data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    ser = QuizBankQuestionSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    q = QuizBankQuestion(**ser.validated_data)
    q.save()
    return Response(QuizBankQuestionSerializer(q).data)


@api_view(["PUT", "DELETE"])
def quiz_bank_detail(request, question_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        q = QuizBankQuestion.objects.get(id=question_id)
    except QuizBankQuestion.DoesNotExist:
        return Response({"detail": "Question not found"}, status=404)

    if request.method == "DELETE":
        q.delete()
        return Response({"ok": True})

    ser = QuizBankQuestionSerializer(data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    for k, v in ser.validated_data.items():
        setattr(q, k, v)
    q.save()
    return Response(QuizBankQuestionSerializer(q).data)


@api_view(["POST"])
def add_bank_question_to_video(request, course_id, video_id, question_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
        bq = QuizBankQuestion.objects.get(id=question_id)
    except (CourseVideo.DoesNotExist, QuizBankQuestion.DoesNotExist):
        return Response({"detail": "Not found"}, status=404)

    if any(_qkey(vq.question) == _qkey(bq.question) for vq in video.quiz_questions.all()):
        return Response({"detail": "Question already on video"}, status=400)

    q = QuizQuestion(
        video=video,
        question=dict(bq.question or {}),
        options=list(bq.options or []),
        correct_index=bq.correct_index,
        correct_indices=list(bq.correct_indices or []),
        fill_answer=dict(bq.fill_answer or {}),
        question_type=bq.question_type or "single",
        order=QuizQuestion.objects.filter(video=video).count(),
    )
    q.save()
    return Response(QuizQuestionSerializer(q).data)


@api_view(["POST"])
def generate_video_quiz(request, course_id, video_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
        course = Course.objects.get(id=course_id)
    except (CourseVideo.DoesNotExist, Course.DoesNotExist):
        return Response({"detail": "Not found"}, status=404)

    count = request.data.get("count") or course.quiz_questions_count or 8
    count = max(1, min(20, int(count)))
    existing_norm = {
        _qkey(qq["question"])
        for qq in QuizQuestion.objects.filter(video=video).values("question")
    }
    pool = list(QuizBankQuestion.objects.all())
    if course.track and course.track != "todas":
        track_pool = [p for p in pool if p.track == course.track]
        if len(track_pool) >= count:
            pool = track_pool
    random.shuffle(pool)
    added = 0
    for bq in pool:
        if added >= count:
            break
        if _qkey(bq.question) in existing_norm:
            continue
        q = QuizQuestion(
            video=video,
            question=dict(bq.question or {}),
            options=list(bq.options or []),
            correct_index=bq.correct_index,
            correct_indices=list(bq.correct_indices or []),
            fill_answer=dict(bq.fill_answer or {}),
            question_type=bq.question_type or "single",
            order=QuizQuestion.objects.filter(video=video).count(),
        )
        q.save()
        existing_norm.add(_qkey(bq.question))
        added += 1
    return Response({"added": added})


# ─── Final Exam ─────────────────────────────────────────

def _qkey(qtext):
    if isinstance(qtext, dict):
        return json.dumps(qtext, sort_keys=True, ensure_ascii=False)
    return str(qtext)


def _exam_questions(course, partner_id=""):
    custom = list(CourseExamQuestion.objects.filter(course=course))
    if custom:
        return [
            {
                "id": q.id,
                "question": q.question,
                "options": q.options,
                "question_type": q.question_type or "single",
                "correct_index": q.correct_index,
                "correct_indices": q.correct_indices or [],
            }
            for q in custom
        ]
    bank = list(QuizBankQuestion.objects.all())
    count = course.exam_questions_count or 5
    if course.track and course.track != "todas":
        track_pool = [p for p in bank if p.track == course.track]
        if len(track_pool) >= count:
            bank = track_pool
    rnd = random.Random(f"{course.id}:{partner_id or 'x'}")
    rnd.shuffle(bank)
    questions = bank[:count]
    used = {_qkey(q.question) for q in questions}
    if len(questions) < count:
        for vq in QuizQuestion.objects.filter(video__course=course):
            if len(questions) >= count:
                break
            if _qkey(vq.question) in used:
                continue
            questions.append({
                "id": f"vq-{vq.id}",
                "question": vq.question,
                "options": vq.options,
                "question_type": vq.question_type or "single",
                "correct_index": vq.correct_index,
                "correct_indices": vq.correct_indices or [],
                "fill_answer": vq.fill_answer or {},
            })
            used.add(_qkey(vq.question))
    return questions


@api_view(["GET"])
def course_exam(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    lang = request.query_params.get("lang", "en")
    data = []
    for q in _exam_questions(course, str(user.id)):
        question = q.get("question") if isinstance(q, dict) else q.question
        options = q.get("options") if isinstance(q, dict) else q.options
        item = {
            "id": q.get("id") if isinstance(q, dict) else q.id,
            "question": (question or {}).get(lang) or (question or {}).get("en") or "" if isinstance(question, dict) else question,
            "question_type": q.get("question_type") if isinstance(q, dict) else (q.question_type or "single"),
        }
        if isinstance(options, list):
            item["options"] = [
                opt.get(lang) or opt.get("en") or "" if isinstance(opt, dict) else opt
                for opt in options
            ]
        else:
            item["options"] = []
        data.append(item)
    return Response(data)


@api_view(["POST"])
def course_exam_submit(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    exam = _exam_questions(course, str(user.id))
    answers = request.data.get("answers", [])
    answers_by_id = {a.get("question_id"): a for a in answers}

    results = []
    correct_count = 0
    for q in exam:
        qid = q.get("id") if isinstance(q, dict) else q.id
        ans = answers_by_id.get(qid)
        if qid.startswith("vq-") and ans is not None:
            real = QuizQuestion.objects.filter(id=qid[3:]).first()
            is_correct = _grade_answer(real, ans) if real else False
        elif ans is not None:
            correct_indices = q.correct_indices or []
            qt = q.question_type or "single"
            if qt == "multiple":
                is_correct = set(ans.get("selected_indices") or []) == set(correct_indices)
            elif qt == "fill":
                given = str(ans.get("answer") or "").strip().lower()
                expected = [str(v).strip().lower() for v in (q.fill_answer or {}).values() if v]
                is_correct = given in expected
            else:
                is_correct = ans.get("selected_index") == q.correct_index
        else:
            is_correct = False
        if is_correct:
            correct_count += 1
        results.append({"question_id": qid, "is_correct": is_correct})

    total = len(exam)
    score = round((correct_count / total) * 100) if total > 0 else 0
    passed = score >= (course.pass_mark or 80)

    prev = TrainingResult.objects.filter(partner=user, course=course).count()
    TrainingResult.objects.create(
        partner=user, course=course, attempt_number=prev + 1,
        score=score, passed=passed,
    )

    if passed:
        prog = PartnerProgress.objects.filter(partner=user, course=course).first()
        if not prog:
            prog = PartnerProgress(partner=user, course=course, completed_videos=[])
        all_vids = list(course.videos.values_list("id", flat=True))
        prog.completed_videos = all_vids
        prog.progress_pct = 100
        prog.completed = True
        prog.score = score
        if not prog.completed_at:
            prog.completed_at = dj_timezone.now()
        prog.save()
        _recompute_certification(user)

    return Response({"score": score, "correct": correct_count, "total": total, "passed": passed, "pass_mark": course.pass_mark, "results": results})


# ─── Custom Exam Questions CRUD ─────────────────────────

@api_view(["GET"])
def exam_question_list(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)
    questions = CourseExamQuestion.objects.filter(course=course)
    data = CourseExamQuestionSerializer(questions, many=True).data
    return Response(data)


@api_view(["POST"])
def exam_question_create(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)
    data = request.data.copy()
    data["course"] = course.id
    ser = CourseExamQuestionSerializer(data=data)
    if not ser.is_valid():
        return Response(ser.errors, status=400)
    order = CourseExamQuestion.objects.filter(course=course).count()
    q = CourseExamQuestion(course=course, order=order, **{k: v for k, v in ser.validated_data.items() if k != "course"})
    q.save()
    return Response(CourseExamQuestionSerializer(q).data, status=201)


@api_view(["PUT", "PATCH"])
def exam_question_update(request, course_id, question_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        q = CourseExamQuestion.objects.get(id=question_id, course_id=course_id)
    except CourseExamQuestion.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    ser = CourseExamQuestionSerializer(data=request.data, partial=True)
    if not ser.is_valid():
        return Response(ser.errors, status=400)
    for k, v in ser.validated_data.items():
        if k != "course":
            setattr(q, k, v)
    q.save()
    return Response(CourseExamQuestionSerializer(q).data)


@api_view(["DELETE"])
def exam_question_delete(request, course_id, question_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        q = CourseExamQuestion.objects.get(id=question_id, course_id=course_id)
    except CourseExamQuestion.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    q.delete()
    return Response({"detail": "Deleted"})


@api_view(["POST"])
def exam_question_reorder(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)
    ids = request.data.get("order", [])
    for i, qid in enumerate(ids):
        CourseExamQuestion.objects.filter(id=qid, course=course).update(order=i)
    return Response({"detail": "ok"})


@api_view(["POST"])
def exam_generate_from_bank(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)
    count = request.data.get("count", course.exam_questions_count or 5)
    bank = list(QuizBankQuestion.objects.all())
    if course.track and course.track != "todas":
        track_pool = [p for p in bank if p.track == course.track]
        if len(track_pool) >= count:
            bank = track_pool
    rnd = random.Random(f"{course.id}:exam")
    rnd.shuffle(bank)
    selected = bank[:count]
    existing = CourseExamQuestion.objects.filter(course=course)
    max_order = existing.count() if existing.exists() else 0
    created = []
    for i, bq in enumerate(selected):
        q = CourseExamQuestion(
            course=course,
            question=bq.question,
            options=bq.options,
            correct_index=bq.correct_index,
            correct_indices=bq.correct_indices or [],
            question_type=bq.question_type or "single",
            order=max_order + i,
        )
        q.save()
        created.append(q)
    data = CourseExamQuestionSerializer(created, many=True).data
    return Response(data, status=201)


# ─── Ratings ────────────────────────────────────────────

@api_view(["GET", "POST"])
def course_rating(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    if request.method == "GET":
        ratings = CourseRating.objects.filter(course=course)
        data = CourseRatingSerializer(ratings, many=True).data
        avg = 0
        if ratings.count():
            avg = round(sum(r.stars for r in ratings) / ratings.count(), 1)
        mine = ratings.filter(partner=user).first()
        return Response({
            "rating_avg": avg,
            "rating_count": ratings.count(),
            "ratings": data,
            "my_rating": CourseRatingSerializer(mine).data if mine else None,
        })

    stars = request.data.get("stars")
    comment = request.data.get("comment", "")
    try:
        stars = int(stars)
    except (TypeError, ValueError):
        return Response({"detail": "Stars must be an integer 1-5"}, status=400)
    if stars < 1 or stars > 5:
        return Response({"detail": "Stars must be between 1 and 5"}, status=400)

    rating, created = CourseRating.objects.update_or_create(
        course=course, partner=user,
        defaults={"stars": stars, "comment": comment},
    )
    return Response(CourseRatingSerializer(rating).data)


# ─── Assignments ────────────────────────────────────────

@api_view(["GET", "POST"])
def course_assignments(request, course_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    if request.method == "GET":
        asgs = CourseAssignment.objects.filter(course=course).select_related("partner")
        return Response(CourseAssignmentSerializer(asgs, many=True).data)

    partner_id = request.data.get("partner_id", "")
    deadline = request.data.get("deadline", "")
    try:
        partner = Partner.objects.get(id=partner_id)
    except Partner.DoesNotExist:
        return Response({"detail": "Partner not found"}, status=404)

    asg, created = CourseAssignment.objects.update_or_create(
        course=course, partner=partner,
        defaults={"deadline": deadline or None},
    )
    if created:
        title = course.title or {}
        if not isinstance(title, dict):
            title = {"en": str(title), "es": str(title), "de": str(title)}
        _notify(
            [partner.id], "curso_asignado", title,
            {"en": "You have been assigned this course", "es": "Te han asignado este curso", "de": "Dieser Kurs wurde Ihnen zugewiesen"},
            "/partner/courses",
        )
    return Response(CourseAssignmentSerializer(asg).data)


@api_view(["DELETE"])
def course_assignment_delete(request, course_id, assignment_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    CourseAssignment.objects.filter(id=assignment_id, course_id=course_id).delete()
    return Response({"ok": True})


# ─── Quiz ──────────────────────────────────────────────

@api_view(["GET", "POST"])
def quiz_questions(request, course_id, video_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
    except CourseVideo.DoesNotExist:
        return Response({"detail": "Video not found"}, status=404)

    lang = request.query_params.get("lang", "en")

    if request.method == "GET":
        questions = QuizQuestion.objects.filter(video=video)
        data = QuizQuestionSerializer(questions, many=True).data
        for q in data:
            if isinstance(q.get("question"), dict):
                q["question"] = q["question"].get(lang) or q["question"].get("en") or ""
            if isinstance(q.get("options"), list):
                q["options"] = [
                    opt.get(lang) or opt.get("en") or "" if isinstance(opt, dict) else opt
                    for opt in q["options"]
                ]
        return Response(data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    ser = QuizQuestionAdminSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    question = QuizQuestion(video=video, **ser.validated_data)
    question.save()
    return Response(QuizQuestionSerializer(question).data)


@api_view(["PUT", "DELETE"])
def quiz_question_detail(request, course_id, video_id, question_id):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    try:
        question = QuizQuestion.objects.get(id=question_id, video_id=video_id)
    except QuizQuestion.DoesNotExist:
        return Response({"detail": "Question not found"}, status=404)

    if request.method == "DELETE":
        question.delete()
        return Response({"ok": True})

    ser = QuizQuestionAdminSerializer(data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    for k, v in ser.validated_data.items():
        setattr(question, k, v)
    question.save()
    return Response(QuizQuestionSerializer(question).data)


def _grade_answer(question, ans):
    """Grade one answer for a question. Supports single/multiple/true_false/fill."""
    qt = question.question_type or "single"
    if qt == "multiple":
        selected = set(ans.get("selected_indices") or [])
        correct = set(question.correct_indices or [])
        return selected == correct
    if qt == "fill":
        given = str(ans.get("answer") or "").strip().lower()
        expected = [str(v).strip().lower() for v in (question.fill_answer or {}).values() if v]
        return given in expected
    selected = ans.get("selected_index")
    return selected == question.correct_index


@api_view(["POST"])
def submit_quiz(request, course_id, video_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
        course = Course.objects.get(id=course_id)
    except (CourseVideo.DoesNotExist, Course.DoesNotExist):
        return Response({"detail": "Video not found"}, status=404)

    answers = request.data.get("answers", [])
    if not answers:
        return Response({"detail": "No answers provided"}, status=400)

    all_questions = list(QuizQuestion.objects.filter(video=video))
    answers_by_id = {a.get("question_id"): a for a in answers}

    results = []
    correct_count = 0
    for question in all_questions:
        ans = answers_by_id.get(question.id)
        is_correct = _grade_answer(question, ans) if ans is not None else False
        if is_correct:
            correct_count += 1
        if ans is not None:
            QuizAttempt.objects.create(
                partner=user, video=video, question=question,
                selected_index=ans.get("selected_index", -1),
                answer_data={k: v for k, v in ans.items() if k != "question_id"},
                is_correct=is_correct,
            )
        results.append({
            "question_id": question.id,
            "is_correct": is_correct,
            "correct_index": question.correct_index,
            "correct_indices": question.correct_indices or [],
            "question_type": question.question_type or "single",
        })

    total = len(all_questions)
    score = round((correct_count / total) * 100) if total > 0 else 0
    pass_mark = course.pass_mark or 80
    passed = score >= pass_mark

    prog = PartnerProgress.objects.filter(partner=user, course_id=course_id).first()
    if not prog:
        prog = PartnerProgress(partner=user, course_id=course_id, completed_videos=[])
    prog.score = score

    completed_vids = list(prog.completed_videos or [])
    if passed and video_id not in completed_vids:
        completed_vids.append(video_id)
    prog.completed_videos = completed_vids

    total_videos = CourseVideo.objects.filter(course_id=course_id).count()
    if total_videos > 0:
        prog.progress_pct = round((len(completed_vids) / total_videos) * 100)
    prog.completed = total_videos > 0 and len(completed_vids) >= total_videos
    if prog.completed and not prog.completed_at:
        prog.completed_at = dj_timezone.now()
    prog.save()

    # Record formal training result (attempt number / passed)
    prev_attempts = TrainingResult.objects.filter(partner=user, course_id=course_id).count()
    if passed:
        TrainingResult.objects.create(
            partner=user, course=course, attempt_number=prev_attempts + 1,
            score=score, passed=True,
        )
    else:
        TrainingResult.objects.create(
            partner=user, course=course, attempt_number=prev_attempts + 1,
            score=score, passed=False,
        )

    if passed and prog.completed:
        _recompute_certification(user)

    return Response({"score": score, "correct": correct_count, "total": total, "passed": passed, "pass_mark": pass_mark, "results": results})


@api_view(["GET"])
def quiz_results(request, course_id, video_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        video = CourseVideo.objects.get(id=video_id, course_id=course_id)
    except CourseVideo.DoesNotExist:
        return Response({"detail": "Video not found"}, status=404)

    attempts = QuizAttempt.objects.filter(partner=user, video=video).select_related("question")
    return Response(QuizAttemptSerializer(attempts, many=True).data)


# ─── Course Progress ────────────────────────────────────

@api_view(["GET"])
def get_partner_progress(request, partner_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    lang = request.query_params.get("lang", "en")
    courses = Course.objects.all()
    result = []
    for c in courses:
        prog = PartnerProgress.objects.filter(partner_id=partner_id, course=c).first()
        latest_result = TrainingResult.objects.filter(partner_id=partner_id, course=c, passed=True).first()
        result.append({
            "course_id": c.id,
            "title": c.title.get(lang) or c.title.get("en") or "",
            "description": c.description.get(lang) or c.description.get("en") or "",
            "thumbnail_url": c.thumbnail_url or "", "category": c.category or "",
            "level": c.level or "beginner", "video_count": c.videos.count(),
            "track": c.track or "", "related_products": c.related_products or [],
            "pass_mark": c.pass_mark, "validity_months": c.validity_months,
            "prerequisite_course_id": c.prerequisite_course_id or "",
            "phase_config": c.phase_config or [],
            "videos": CourseVideoSerializer(c.videos.all(), many=True, lang=lang).data,
            "progress_pct": prog.progress_pct if prog else 0,
            "completed": prog.completed if prog else False,
            "completed_videos": prog.completed_videos if prog else [],
            "certificate_url": latest_result.certificate_url if latest_result else "",
        })
    return Response(result)


@api_view(["GET"])
def get_course_video_progress(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    prog = PartnerProgress.objects.filter(partner=user, course_id=course_id).first()
    completed = prog.completed_videos if prog else []
    return Response({vid: True for vid in completed})


@api_view(["POST"])
def update_progress(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    course_id = request.data.get("course_id")
    video_id = request.data.get("video_id")
    completed_flag = request.data.get("completed", False)

    prog = PartnerProgress.objects.filter(partner=user, course_id=course_id).first()
    if not prog:
        prog = PartnerProgress(partner=user, course_id=course_id, completed_videos=[])

    completed_vids = list(prog.completed_videos or [])

    if video_id:
        if completed_flag and video_id not in completed_vids:
            completed_vids.append(video_id)
        elif not completed_flag and video_id in completed_vids:
            completed_vids.remove(video_id)

    prog.completed_videos = completed_vids

    total_videos = CourseVideo.objects.filter(course_id=course_id).count()
    if total_videos > 0:
        prog.progress_pct = round((len(completed_vids) / total_videos) * 100)
    prog.completed = total_videos > 0 and len(completed_vids) >= total_videos
    if prog.completed and not prog.completed_at:
        prog.completed_at = dj_timezone.now()

    prog.save()
    return Response({"ok": True, "completed_videos": completed_vids, "progress_pct": prog.progress_pct, "completed": prog.completed})


# ─── Deals ──────────────────────────────────────────────

@api_view(["GET", "POST"])
def deals_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        deals = Deal.objects.filter(partner=user).select_related("partner").order_by("-created_at")
        return Response(DealSerializer(deals, many=True).data)

    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados pueden crear deals"}, status=403)
    ser = DealCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    deal = Deal(partner=user, **ser.validated_data)
    deal.save()
    return Response({"id": deal.id})


@api_view(["GET"])
def list_all_deals(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    deals = Deal.objects.select_related("partner").order_by("-created_at")
    return Response(DealSerializer(deals, many=True).data)


@api_view(["PATCH", "DELETE"])
def deal_detail(request, deal_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        deal = Deal.objects.select_related("partner").get(id=deal_id)
    except Deal.DoesNotExist:
        return Response({"detail": "Deal not found"}, status=404)

    if user.role == "socio" and deal.partner_id != user.id:
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "DELETE":
        Commission.objects.filter(deal=deal).delete()
        deal.delete()
        return Response({"ok": True})

    ser = DealUpdateSerializer(data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    old_status = deal.status
    for k, v in ser.validated_data.items():
        setattr(deal, k, v)

    if ser.validated_data.get("status") == "completado" and not hasattr(deal, "commission"):
        amount = deal.estimated_value * (deal.partner.commission_rate / 100)
        Commission.objects.create(
            partner=deal.partner, deal=deal,
            amount=round(amount, 2), status="pendiente",
        )
    deal.save()
    if ser.validated_data.get("status") and ser.validated_data.get("status") != old_status:
        labels = {
            "necesita_acceso": {"en": "Needs access", "es": "Necesita acceso", "de": "Zugriff erforderlich"},
            "en_revision": {"en": "Under review", "es": "En revisión", "de": "In Prüfung"},
            "en_implementacion": {"en": "In implementation", "es": "En implementación", "de": "In Implementierung"},
            "acceso_otorgado": {"en": "Access granted", "es": "Acceso otorgado", "de": "Zugriff gewährt"},
            "completado": {"en": "Completed", "es": "Completado", "de": "Abgeschlossen"},
            "perdido": {"en": "Lost", "es": "Perdido", "de": "Verloren"},
        }
        st = labels.get(deal.status, {"en": deal.status, "es": deal.status, "de": deal.status})
        _notify(
            [deal.partner_id], "deal",
            {"en": f"Deal updated: {deal.company_name}", "es": f"Deal actualizado: {deal.company_name}", "de": f"Deal aktualisiert: {deal.company_name}"},
            {"en": f"Status changed to {st['en']}", "es": f"El estado cambió a {st['es']}", "de": f"Status geändert zu {st['de']}"},
            "/partner/deals",
        )
    return Response({"ok": True})


# ─── Reports ────────────────────────────────────────────

@api_view(["GET"])
def admin_stats(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    non_admin = Partner.objects.exclude(role="admin")
    total_partners = non_admin.count()
    active_partners = non_admin.filter(status="activo").count()
    pending_requests = non_admin.filter(status__in=["solicitado", "en_revision"]).count()
    total_deals = Deal.objects.count()
    total_revenue = Deal.objects.filter(status="completado").aggregate(s=Sum("estimated_value"))["s"] or 0
    pending_commissions = Commission.objects.filter(status="pendiente").aggregate(s=Sum("amount"))["s"] or 0

    # Pipeline stats
    opps = Opportunity.objects.all()
    total_pipeline_value = opps.exclude(stage="perdida").aggregate(s=Sum("amount"))["s"] or 0
    weighted_pipeline_value = sum(
        (o.amount or 0) * (o.probability or 0) / 100
        for o in opps.exclude(stage="perdida")
    )
    active_opportunities = opps.exclude(stage="perdida").count()
    by_stage = {}
    for o in opps:
        by_stage.setdefault(o.stage, {"count": 0, "value": 0})
        by_stage[o.stage]["count"] += 1
        by_stage[o.stage]["value"] += o.amount or 0

    # Security stats
    from datetime import timedelta
    from django.utils import timezone
    last_24h = timezone.now() - timedelta(hours=24)
    failed_logins_24h = LoginAttempt.objects.filter(success=False, attempted_at__gte=last_24h).count()
    blacklisted_tokens_count = TokenBlacklist.objects.count()

    top = []
    for p in non_admin.filter(status="activo"):
        deal_count = Deal.objects.filter(partner=p).count()
        deal_revenue = Deal.objects.filter(partner=p, status="completado").aggregate(s=Sum("estimated_value"))["s"] or 0
        top.append({"id": p.id, "name": p.company_name, "deals": deal_count, "revenue": deal_revenue})
    top.sort(key=lambda x: x["revenue"], reverse=True)

    deals_by_status = {}
    for d in Deal.objects.all():
        deals_by_status[d.status] = deals_by_status.get(d.status, 0) + 1

    return Response({
        "total_partners": total_partners, "active_partners": active_partners,
        "pending_requests": pending_requests, "total_deals": total_deals,
        "total_revenue": total_revenue, "pending_commissions": pending_commissions,
        "total_pipeline_value": total_pipeline_value,
        "weighted_pipeline_value": round(weighted_pipeline_value, 2),
        "active_opportunities": active_opportunities,
        "pipeline_by_stage": by_stage,
        "deals_by_status": deals_by_status,
        "failed_logins_24h": failed_logins_24h,
        "blacklisted_tokens_count": blacklisted_tokens_count,
        "top_partners": top[:10],
    })


@api_view(["GET"])
def partner_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    active_deals = Deal.objects.filter(partner=user).exclude(status__in=["completado", "perdido"]).count()
    completed_deals = Deal.objects.filter(partner=user, status="completado").count()
    total_revenue = Deal.objects.filter(partner=user, status="completado").aggregate(s=Sum("estimated_value"))["s"] or 0
    commissions_earned = Commission.objects.filter(partner=user, status="pagada").aggregate(s=Sum("amount"))["s"] or 0
    pending_commissions = Commission.objects.filter(partner=user, status="pendiente").aggregate(s=Sum("amount"))["s"] or 0
    total_courses = Course.objects.count()
    courses_completed = PartnerProgress.objects.filter(partner=user, completed=True).count()
    completion_rate = round((courses_completed / total_courses * 100)) if total_courses > 0 else 0
    courses_enrolled = PartnerProgress.objects.filter(partner=user).values("course").distinct().count()

    status_breakdown = {}
    for s in ["necesita_acceso", "en_revision", "en_implementacion", "acceso_otorgado", "completado"]:
        status_breakdown[s] = Deal.objects.filter(partner=user, status=s).count()

    return Response({
        "active_deals": active_deals, "completed_deals": completed_deals,
        "total_revenue": total_revenue, "commissions_earned": commissions_earned,
        "pending_commissions": pending_commissions,
        "courses_enrolled": courses_enrolled, "completion_rate": completion_rate,
        "status_breakdown": status_breakdown,
    })


# ─── Partner Commissions ────────────────────────────────

@api_view(["GET"])
def my_commissions(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    commissions = Commission.objects.filter(partner=user).select_related("deal").order_by("-created_at")
    return Response(CommissionSerializer(commissions, many=True).data)


# ─── Notifications ──────────────────────────────────────

def _notify(partner_ids, ntype, title, message, link=""):
    """Create a notification for each partner. title/message are {en, es, de} dicts."""
    ids = list(dict.fromkeys(str(p) for p in partner_ids if p))
    if not ids:
        return 0
    objs = [
        Notification(partner_id=pid, type=ntype, title=title, message=message, link=link)
        for pid in ids
    ]
    Notification.objects.bulk_create(objs)
    return len(objs)


def _localized(en, es=None, de=None):
    return {"en": en, "es": es or en, "de": de or en}


def _notify_all_active(ntype, title, message, link=""):
    partners = Partner.objects.filter(status="activo").exclude(role="admin")
    return _notify([p.id for p in partners], ntype, title, message, link)


@api_view(["GET"])
def notifications_list(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    lang = request.query_params.get("lang", "en")
    qs = Notification.objects.filter(partner=user)
    return Response(NotificationSerializer(qs[:50], many=True, context={"lang": lang}).data)


@api_view(["GET"])
def notifications_unread_count(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    count = Notification.objects.filter(partner=user, read=False).count()
    return Response({"count": count})


@api_view(["PATCH"])
def notification_mark_read(request, notification_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    Notification.objects.filter(id=notification_id, partner=user).update(read=True)
    return Response({"ok": True})


@api_view(["POST"])
def notifications_mark_all_read(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    Notification.objects.filter(partner=user, read=False).update(read=True)
    return Response({"ok": True})


@api_view(["POST"])
def notifications_broadcast(request):
    """Admin sends an announcement to all active partners."""
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    title = request.data.get("title")
    message = request.data.get("message") or ""
    if not title or not str(title).strip():
        return Response({"detail": "Título obligatorio"}, status=400)
    title = _localized(str(title).strip())
    message = _localized(str(message).strip())
    link = (request.data.get("link") or "").strip()
    count = _notify_all_active("anuncio", title, message, link)
    return Response({"ok": True, "sent": count})


# ─── Product Catalog ────────────────────────────────────

def _normalize_name(value):
    """Accept a string (applied to all langs) or a {en, es, de} dict."""
    if isinstance(value, dict):
        return {k: str(v or "").strip() for k, v in value.items() if k in ("en", "es", "de")} or {"en": ""}
    s = str(value or "").strip()
    return {"en": s, "es": s, "de": s}


@api_view(["GET", "POST"])
def products_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        qs = Product.objects.all()
        if user.role != "admin":
            qs = qs.filter(active=True)
        return Response(ProductSerializer(qs, many=True).data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    key = (request.data.get("key") or "").strip().lower().replace(" ", "_")
    if not key:
        return Response({"detail": "Key del producto obligatorio"}, status=400)
    if Product.objects.filter(key=key).exists():
        return Response({"detail": "El producto ya existe"}, status=400)

    prod = Product(
        key=key,
        name=_normalize_name(request.data.get("name") or key),
        description=_normalize_name(request.data.get("description") or ""),
        price_usd=request.data.get("price_usd") or 0,
        price_eur=request.data.get("price_eur") or 0,
        price_chf=request.data.get("price_chf") or 0,
        price_otro=request.data.get("price_otro") or 0,
        custom_currency=(request.data.get("custom_currency") or "").strip(),
        category=(request.data.get("category") or "core").strip() or "core",
        active=bool(request.data.get("active", True)),
        sort_order=int(request.data.get("sort_order") or 0),
    )
    prod.save()
    return Response(ProductSerializer(prod).data, status=201)


@api_view(["GET", "PATCH", "DELETE"])
def product_detail(request, product_key):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        prod = Product.objects.get(key=product_key)
    except Product.DoesNotExist:
        return Response({"detail": "Product not found"}, status=404)

    if request.method == "GET":
        return Response(ProductSerializer(prod).data)

    if user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    if request.method == "DELETE":
        prod.delete()
        return Response({"ok": True})

    d = request.data
    if "name" in d:
        prod.name = _normalize_name(d["name"])
    if "description" in d:
        prod.description = _normalize_name(d["description"])
    if "price_usd" in d:
        prod.price_usd = d["price_usd"] or 0
    if "price_eur" in d:
        prod.price_eur = d["price_eur"] or 0
    if "price_chf" in d:
        prod.price_chf = d["price_chf"] or 0
    if "price_otro" in d:
        prod.price_otro = d["price_otro"] or 0
    if "custom_currency" in d:
        prod.custom_currency = (d["custom_currency"] or "").strip()
    if "category" in d:
        prod.category = (d["category"] or "core").strip() or "core"
    if "active" in d:
        prod.active = bool(d["active"])
    if "sort_order" in d:
        prod.sort_order = int(d["sort_order"] or 0)
    prod.save()
    return Response(ProductSerializer(prod).data)


# ─── Chatbot ────────────────────────────────────────────

@api_view(["POST"])
def chat_message(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    raw = request.data.get("messages")
    if not isinstance(raw, list) or not raw:
        return Response({"detail": "messages requerido"}, status=400)

    clean = []
    for m in raw[-20:]:
        if not isinstance(m, dict):
            continue
        role = m.get("role")
        content = str(m.get("content") or "").strip()
        if role in ("user", "assistant") and content:
            clean.append({"role": role, "content": content[:4000]})

    while clean and clean[0]["role"] != "user":
        clean.pop(0)
    if not clean:
        return Response({"detail": "messages vacío"}, status=400)

    from .ai import ask_ai
    reply, err = ask_ai(clean)
    if err:
        return Response({"detail": err}, status=503)
    return Response({"reply": reply})


# ─── Pipeline ───────────────────────────────────────────

def _ensure_conflict_record(opp, other):
    existing = ChannelConflict.objects.filter(
        company_name__iexact=opp.company_name
    ).exclude(status__in=["resuelto", "cerrado"]).first()
    if existing:
        fields = []
        if not existing.opportunity:
            existing.opportunity = opp
            fields.append("opportunity")
        if not existing.conflicting_opportunity and other:
            existing.conflicting_opportunity = other
            fields.append("conflicting_opportunity")
        if fields:
            existing.save(update_fields=fields)
        return existing
    return ChannelConflict.objects.create(
        company_name=opp.company_name,
        opportunity=opp,
        conflicting_opportunity=other,
        reporter=opp.partner,
        notes="Detectado automáticamente por coincidencia de empresa",
    )


def _close_conflicts_for_opportunity(opp, reason):
    qs = ChannelConflict.objects.filter(
        Q(opportunity=opp) | Q(conflicting_opportunity=opp)
    ).exclude(status__in=["resuelto", "cerrado"])
    now = dj_timezone.now()
    for c in qs:
        c.status = "cerrado"
        c.resolution = reason
        c.resolved_at = now
        c.save(update_fields=["status", "resolution", "resolved_at"])
        for o in (c.opportunity, c.conflicting_opportunity):
            if o and o.conflict_indicator:
                o.conflict_indicator = False
                o.save(update_fields=["conflict_indicator"])


def _detect_conflict(opp):
    """Mark conflict on this and all active opportunities of the same company."""
    dup = Opportunity.objects.filter(company_name__iexact=opp.company_name).exclude(pk=opp.pk).exclude(stage="perdida")
    if dup.exists():
        opp.conflict_indicator = True
        opp.save(update_fields=["conflict_indicator"])
        dup.update(conflict_indicator=True)
        _ensure_conflict_record(opp, dup.first())
        return True
    return False


def _progressive_validation_errors(opp):
    errs = []
    if not (opp.company_name or "").strip():
        errs.append("Empresa obligatoria")
    if not opp.company_size:
        errs.append("Tamaño de empresa obligatorio")
    if not opp.products:
        errs.append("Producto(s) obligatorio(s)")
    if not opp.operation_mode:
        errs.append("Modalidad de operación obligatoria")
    if not (opp.delivery_quarter or "").strip():
        errs.append("Trimestre de entrega obligatorio")
    if opp.stage in ("propuesta_enviada", "negociacion", "ganada"):
        if not opp.amount:
            errs.append("Amount (ARR) obligatorio desde Propuesta enviada")
        if not opp.forecast_category:
            errs.append("Forecast Category obligatoria desde Propuesta enviada")
    return errs


def _refresh_certification_status(partner):
    now = dj_timezone.now()
    for cert in Certification.objects.filter(partner=partner):
        if cert.valid_until and cert.valid_until < now:
            cert.status = "expired"
        else:
            old_completed = PartnerProgress.objects.filter(
                partner=partner, completed=True,
                completed_at__lt=now - timedelta(days=365),
            ).exists()
            cert.status = "pending_update" if old_completed else "valid"
        cert.save(update_fields=["status"])


def _recompute_certification(partner):
    """Programa de 5 fases: al completar el curso con nota mayor a 80, el partner se certifica."""
    best = (
        TrainingResult.objects.filter(partner=partner, passed=True)
        .select_related("course")
        .order_by("-score")
        .first()
    )
    if not best or best.score <= 80:
        return None

    prog = PartnerProgress.objects.filter(
        partner=partner, course=best.course, completed=True
    ).first()
    if not prog:
        return None

    validity_months = best.course.validity_months or 12
    valid_until = dj_timezone.now() + timedelta(days=30 * validity_months)

    cert = Certification.objects.filter(partner=partner).first()
    was_new = cert is None
    if not cert:
        cert = Certification(partner=partner, level="associate")
    cert.status = "valid"
    cert.valid_until = valid_until
    cert.save()

    if was_new:
        award_points(
            partner, "certificacion",
            note=f"Nota: {best.score}",
            source_type="cert", source_id=cert.id,
        )
    return cert


_SWEEP_INTERVAL_HOURS = 6
_sweep_last_run = None


def _run_pipeline_sweep():
    """Degrada oportunidades estancadas y reporta las que están por vencer."""
    now = dj_timezone.now()
    stale_45 = []
    demoted = []
    for opp in Opportunity.objects.exclude(stage__in=["ganada", "perdida"]):
        days_inactive = (now - opp.last_activity).days
        if days_inactive > 90:
            from_stage = opp.stage
            opp.stage = "registrada"
            opp.probability = Opportunity.STAGE_PROBABILITY["registrada"]
            opp.last_activity = now
            opp.save()
            OpportunityEvent.objects.create(
                opportunity=opp, from_stage=from_stage, to_stage="registrada",
                field_changes={"auto": "demoted_stale"},
            )
            demoted.append(opp.id)
        elif days_inactive > 45:
            stale_45.append({"id": opp.id, "company_name": opp.company_name, "days": days_inactive})

    conflicts = Opportunity.objects.filter(conflict_indicator=True).values_list("id", flat=True)
    return {
        "stale_over_45_days": stale_45,
        "demoted_to_registrada": demoted,
        "conflicts": list(conflicts),
    }


def _maybe_auto_sweep():
    """Ejecuta el sweep automáticamente como mucho cada _SWEEP_INTERVAL_HOURS."""
    global _sweep_last_run
    now = dj_timezone.now()
    if _sweep_last_run is None or (now - _sweep_last_run).total_seconds() >= _SWEEP_INTERVAL_HOURS * 3600:
        _sweep_last_run = now
        _run_pipeline_sweep()


@api_view(["GET", "POST"])
def pipeline_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        _maybe_auto_sweep()
        if user.role == "admin":
            opps = Opportunity.objects.select_related("partner").all()
        else:
            opps = Opportunity.objects.filter(partner=user)
        stage = request.query_params.get("stage")
        if stage:
            opps = opps.filter(stage=stage)
        search = request.query_params.get("search")
        if search:
            opps = opps.filter(Q(company_name__icontains=search) | Q(name__icontains=search))
        return Response(OpportunitySerializer(opps, many=True).data)

    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    ser = OpportunityCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data
    stage = d.get("stage", "registrada")
    probability = d.get("probability") or Opportunity.STAGE_PROBABILITY.get(stage, 10)

    opp = Opportunity(
        partner=user,
        company_name=d["company_name"],
        company_size=d.get("company_size", "<250"),
        products=d.get("products", []),
        operation_mode=d.get("operation_mode", "cloud"),
        stage=stage,
        probability=probability,
        amount=d.get("amount", 0.0),
        scan_one_time_fee=d.get("scan_one_time_fee", 0.0),
        currency=(d.get("currency", "usd") if user.role == "admin" else currency_for_country(user.country)),
        custom_currency=d.get("custom_currency", ""),
        delivery_quarter=d.get("delivery_quarter", ""),
        close_date=d.get("close_date"),
        deal_owner=d.get("deal_owner", ""),
        channel_manager=d.get("channel_manager", ""),
        opportunity_type=d.get("opportunity_type", "nuevo"),
        forecast_category=d.get("forecast_category"),
        lead_source=d.get("lead_source", "generada_partner"),
        next_steps=d.get("next_steps", ""),
        notes=d.get("notes", ""),
        protection_end_date=dj_timezone.localdate() + timedelta(days=90),
    )
    opp.save()

    errs = _progressive_validation_errors(opp)
    if errs:
        opp.delete()
        return Response({"detail": " · ".join(errs)}, status=400)

    _detect_conflict(opp)
    OpportunityEvent.objects.create(opportunity=opp, to_stage=stage)
    return Response(OpportunitySerializer(opp).data, status=201)


@api_view(["GET", "PATCH", "DELETE"])
def opportunity_detail(request, opp_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        opp = Opportunity.objects.select_related("partner").get(id=opp_id)
    except Opportunity.DoesNotExist:
        return Response({"detail": "Opportunity not found"}, status=404)

    if user.role == "socio" and opp.partner_id != user.id:
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "GET":
        return Response(OpportunitySerializer(opp).data)

    if request.method == "DELETE":
        if user.role != "admin":
            return Response({"detail": "Admin access required"}, status=403)
        _close_conflicts_for_opportunity(opp, "Oportunidad eliminada")
        opp.delete()
        return Response({"ok": True})

    ser = OpportunityUpdateSerializer(data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data
    if user.role == "socio":
        d["currency"] = currency_for_country(user.country)
    from_stage = opp.stage
    to_stage = d.get("stage", opp.stage)

    if from_stage != to_stage:
        errs = []
        if to_stage == "perdida" and not d.get("loss_reason"):
            errs.append("Motivo de pérdida obligatorio")
        if to_stage == "ganada" and not d.get("amount", opp.amount):
            errs.append("Amount (ARR) obligatorio para marcar como ganada")
        if not (d.get("next_steps") or opp.next_steps or "").strip():
            errs.append("Próximos pasos obligatorios al cambiar de etapa")
        if errs:
            return Response({"detail": " · ".join(errs)}, status=400)

    if to_stage == "perdida":
        d["probability"] = 0
    if to_stage == "ganada":
        d["probability"] = 100

    changed = {}
    for k, v in d.items():
        old = getattr(opp, k, None)
        if old != v:
            changed[k] = v
        setattr(opp, k, v)
    if "stage" in d and "probability" not in d:
        opp.probability = Opportunity.STAGE_PROBABILITY.get(d["stage"], opp.probability)

    opp.last_activity = dj_timezone.now()
    opp.save()
    _detect_conflict(opp)

    if to_stage == "perdida":
        _close_conflicts_for_opportunity(opp, "Oportunidad marcada como perdida")

    errs = _progressive_validation_errors(opp)
    if errs:
        return Response({"detail": " · ".join(errs)}, status=400)

    if from_stage != to_stage:
        OpportunityEvent.objects.create(opportunity=opp, from_stage=from_stage, to_stage=to_stage, field_changes=changed)
    if to_stage == "ganada":
        award_points(opp.partner, "deal_ganado", note=opp.name, source_type="opportunity", source_id=opp.id)
        _recompute_certification(opp.partner)
    return Response(OpportunitySerializer(opp).data)


@api_view(["GET"])
def opportunity_events(request, opp_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        opp = Opportunity.objects.get(id=opp_id)
    except Opportunity.DoesNotExist:
        return Response({"detail": "Opportunity not found"}, status=404)
    if user.role == "socio" and opp.partner_id != user.id:
        return Response({"detail": "Access denied"}, status=403)
    events = OpportunityEvent.objects.filter(opportunity=opp)
    return Response(OpportunityEventSerializer(events, many=True).data)


@api_view(["GET"])
def pipeline_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    _maybe_auto_sweep()

    if user.role == "admin":
        opps = Opportunity.objects.all()
    else:
        opps = Opportunity.objects.filter(partner=user)

    total_value = opps.exclude(stage="perdida").aggregate(s=Sum("amount"))["s"] or 0
    weighted_value = opps.exclude(stage="perdida").aggregate(
        s=Sum(F("amount") * F("probability") / 100.0)
    )["s"] or 0
    avg_probability = opps.exclude(stage="perdida").aggregate(
        s=Avg("probability")
    )["s"] or 0

    by_stage = {}
    for stage_key, _ in Opportunity.STAGE_CHOICES:
        stage_opps = opps.filter(stage=stage_key)
        by_stage[stage_key] = {
            "count": stage_opps.count(),
            "value": stage_opps.aggregate(s=Sum("amount"))["s"] or 0,
            "probability": Opportunity.STAGE_PROBABILITY.get(stage_key, 0),
        }

    return Response({
        "total_value": round(total_value, 2),
        "weighted_value": round(weighted_value, 2),
        "avg_probability": round(avg_probability),
        "total_opportunities": opps.count(),
        "conflicts": opps.filter(conflict_indicator=True).count(),
        "by_stage": by_stage,
    })


# ─── Pipeline Automation ────────────────────────────────

@api_view(["POST"])
def pipeline_automation(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    return Response(_run_pipeline_sweep())


# ─── LMS ────────────────────────────────────────────────

@api_view(["GET"])
def lms_overview(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    lang = request.query_params.get("lang", "en")
    _refresh_certification_status(user)
    certs = Certification.objects.filter(partner=user)
    results = TrainingResult.objects.filter(partner=user).select_related("course").order_by("-created_at")

    courses = Course.objects.all()
    course_map = {c.id: c for c in courses}
    completed = set(PartnerProgress.objects.filter(partner=user, completed=True).values_list("course_id", flat=True))

    def track_status(track):
        ids = [c.id for c in course_map.values() if c.track == track]
        done = sum(1 for i in ids if i in completed)
        return {"total": len(ids), "completed": done}

    current_cert = certs.first()
    current_level = current_cert.level if current_cert else None
    tracks = {
        "ventas": track_status("ventas"),
        "tecnica": track_status("tecnica"),
        "cumplimiento": track_status("cumplimiento"),
    }
    won_count = Opportunity.objects.filter(partner=user, stage="ganada").count()

    def _all_done(track):
        st = tracks[track]
        return st["total"] > 0 and st["completed"] >= st["total"]

    if current_level == "professional":
        next_level = {
            "level": "expert",
            "requirements": [
                {"key": "ventas", "label": "Track Ventas", "done": _all_done("ventas")},
                {"key": "tecnica", "label": "Track Técnica", "done": _all_done("tecnica")},
                {"key": "won", "label": "2 oportunidades ganadas", "done": won_count >= 2},
            ],
        }
    elif current_level == "associate":
        next_level = {
            "level": "professional",
            "requirements": [
                {"key": "ventas", "label": "Track Ventas", "done": _all_done("ventas")},
                {"key": "tecnica", "label": "Track Técnica", "done": _all_done("tecnica")},
            ],
        }
    else:
        next_level = {
            "level": "associate",
            "requirements": [
                {"key": "cumplimiento", "label": "Track Compliance", "done": _all_done("cumplimiento")},
                {"key": "producto", "label": "1 curso de producto", "done": tracks["ventas"]["completed"] > 0 or tracks["tecnica"]["completed"] > 0},
            ],
        }

    return Response({
        "certifications": CertificationSerializer(certs, many=True).data,
        "training_results": TrainingResultSerializer(results, many=True).data,
        "tracks": tracks,
        "completed_courses": len(completed),
        "won_opportunities": won_count,
        "next_level": next_level,
    })


@api_view(["GET"])
def lms_report(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    courses = []
    for c in Course.objects.all():
        started = PartnerProgress.objects.filter(course=c).exclude(progress_pct=0).count()
        completed = PartnerProgress.objects.filter(course=c, completed=True).count()
        passed_results = TrainingResult.objects.filter(course=c, passed=True).count()
        total_results = TrainingResult.objects.filter(course=c).count()
        courses.append({
            "course_id": c.id,
            "title": c.title.get("en") or "",
            "track": c.track or "",
            "started": started,
            "completed": completed,
            "completion_rate": round(completed / started * 100) if started else 0,
            "pass_rate": round(passed_results / total_results * 100) if total_results else 0,
            "pass_mark": c.pass_mark,
            "validity_months": c.validity_months,
        })

    partners = []
    for p in Partner.objects.exclude(role="admin"):
        _refresh_certification_status(p)
        cert = Certification.objects.filter(partner=p).first()
        partners.append({
            "partner_id": p.id,
            "company_name": p.company_name,
            "track": p.training_track or "",
            "courses_completed": PartnerProgress.objects.filter(partner=p, completed=True).count(),
            "certification": cert.level if cert else None,
            "cert_status": cert.status if cert else None,
        })

    return Response({"courses": courses, "partners": partners})


@api_view(["GET"])
def certificate_download(request, course_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return Response({"detail": "Course not found"}, status=404)

    prog = PartnerProgress.objects.filter(partner=user, course=course, completed=True).first()
    if not prog:
        return Response({"detail": "Curso no completado"}, status=400)

    from .certificate import build_certificate_pdf
    lang = request.query_params.get("lang", "es")
    date_str = dj_timezone.now().strftime("%d/%m/%Y")
    pdf_bytes = build_certificate_pdf(
        partner_name=user.company_name,
        course_title=course.title.get(lang) or course.title.get("en") or "",
        score=prog.score or 100,
        date_str=date_str,
        lang=lang,
    )

    cert_dir = os.path.join(settings.MEDIA_ROOT, "certificates")
    os.makedirs(cert_dir, exist_ok=True)
    filename = f"cert_{course.id}_{user.id}.pdf"
    filepath = os.path.join(cert_dir, filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    TrainingResult.objects.filter(partner=user, course=course, passed=True).update(
        certificate_url=f"/uploads/certificates/{filename}"
    )
    return Response({"certificate_url": f"/uploads/certificates/{filename}"}, status=201)


# ─── Health ─────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([AllowAny])
def health(request):
    return Response({"status": "ok", "app": "aconso Partner Academy"})


# ─── Admin Security Panel ──────────────────────────────


@api_view(["GET"])
def admin_login_attempts(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)
    limit = int(request.GET.get("limit", "100"))
    attempts = LoginAttempt.objects.all().order_by("-attempted_at")[:limit]
    return Response([
        {
            "ip": a.ip_address,
            "attempted_at": a.attempted_at.isoformat(),
            "success": a.success,
        }
        for a in attempts
    ])


@api_view(["GET", "DELETE"])
def admin_blacklisted_tokens(request, jti=None):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    if request.method == "GET":
        tokens = TokenBlacklist.objects.all().order_by("-created_at")[:100]
        return Response([
            {
                "jti": t.jti,
                "user_id": t.user_id,
                "user_email": t.user.email if hasattr(t.user, "email") else None,
                "created_at": t.created_at.isoformat(),
            }
            for t in tokens
        ])

    if request.method == "DELETE" and jti:
        TokenBlacklist.objects.filter(jti=jti).delete()
        return Response({"detail": "Token removed from blacklist"}, status=200)


@api_view(["POST"])
def admin_register_admin(request):
    user = _current_user(request)
    if not user or user.role != "admin":
        return Response({"detail": "Admin access required"}, status=403)

    ser = PartnerCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data

    if Partner.objects.filter(email=d["email"]).exists():
        return Response({"detail": "El email ya esta registrado"}, status=400)

    password = d.get("password", "")
    password_err = security.validate_password_strength(password)
    if password_err:
        return Response({"detail": password_err}, status=400)
    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

    admin_user = Partner(
        company_name=d["company_name"],
        email=d["email"],
        password_hash=password_hash,
        phone=d.get("phone", ""),
        tax_id=d.get("tax_id", ""),
        contact_name=d.get("contact_name", ""),
        role="admin",
        status="activo",
    )
    admin_user.save()
    return Response(PartnerSerializer(admin_user).data, status=201)


# ─── Onboarding Checklist (RQ-13) ───────────────────────


@api_view(["GET", "POST"])
def onboarding_view(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    lang = request.query_params.get("lang", "en")

    if request.method == "POST":
        if not PartnerOnboarding.objects.filter(partner=user).exists():
            PartnerOnboarding.objects.create(partner=user)
        return Response(onboarding_snapshot(user, lang))

    return Response(onboarding_snapshot(user, lang))


@api_view(["PATCH"])
def onboarding_step(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    key = request.data.get("key", "")
    done = bool(request.data.get("done", False))
    if key not in STEP_ORDER:
        return Response({"detail": "Unknown step"}, status=400)

    lang = request.query_params.get("lang", "en")
    ob, _ = PartnerOnboarding.objects.get_or_create(partner=user)
    steps = list(ob.manual_steps or [])
    if done and key not in steps:
        steps.append(key)
    elif not done and key in steps:
        steps.remove(key)
    ob.manual_steps = steps
    ob.save()
    return Response(onboarding_snapshot(user, lang))


# ─── Partner Users (RQ-01) ──────────────────────────────


def _serialize_member(m, with_token=False):
    d = {
        "id": m.id,
        "partner_id": m.partner_id,
        "email": m.email,
        "contact_name": m.contact_name,
        "role": m.role,
        "status": m.status,
        "created_at": m.created_at.isoformat() if m.created_at else "",
    }
    if with_token:
        d["invite_token"] = m.invite_token
        d["invite_url"] = f"/register?invite={m.invite_token}"
    return d


def _can_manage_members(user):
    return user and user.role == "admin" or (user and user.role == "socio" and not getattr(user, "is_member", False))


@api_view(["GET", "POST"])
def partner_users_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role not in ("admin", "socio"):
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "GET":
        if user.role == "admin":
            qs = PartnerUser.objects.select_related("partner")
            pid = request.query_params.get("partner_id")
            if pid:
                qs = qs.filter(partner_id=pid)
        else:
            qs = PartnerUser.objects.filter(partner=user)
        return Response([_serialize_member(m) for m in qs])

    if not _can_manage_members(user):
        return Response({"detail": "Solo el administrador del partner puede invitar"}, status=403)

    email = (request.data.get("email") or "").strip().lower()
    contact_name = (request.data.get("contact_name") or "").strip()
    role = request.data.get("role") or "member"
    if role not in ("admin", "member"):
        role = "member"
    if not email:
        return Response({"detail": "Email obligatorio"}, status=400)
    if Partner.objects.filter(email=email).exists() or PartnerUser.objects.filter(email=email).exists():
        return Response({"detail": "El email ya está registrado"}, status=400)

    invite_token = uuid.uuid4().hex
    m = PartnerUser.objects.create(
        partner_id=(user.id if user.role == "socio" else request.data.get("partner_id") or user.id),
        email=email, contact_name=contact_name, role=role,
        status="invitado", invite_token=invite_token,
    )
    return Response(_serialize_member(m, with_token=True), status=201)


@api_view(["GET"])
@permission_classes([AllowAny])
def partner_user_invite_info(request, token):
    m = PartnerUser.objects.filter(invite_token=token).select_related("partner").first()
    if not m:
        return Response({"detail": "Invitación no válida"}, status=404)
    return Response({
        "valid": True,
        "email": m.email,
        "partner_name": m.partner.company_name,
        "contact_name": m.contact_name,
        "invite_token": m.invite_token,
    })


@api_view(["POST"])
@permission_classes([AllowAny])
def partner_user_register(request):
    token = request.data.get("invite_token", "")
    contact_name = (request.data.get("contact_name") or "").strip()
    password = request.data.get("password", "")

    m = PartnerUser.objects.filter(invite_token=token).first()
    if not m:
        return Response({"detail": "Invitación no válida"}, status=404)
    if m.status == "activo":
        return Response({"detail": "Esta invitación ya fue utilizada"}, status=400)

    password_err = security.validate_password_strength(password)
    if password_err:
        return Response({"detail": password_err}, status=400)
    if not contact_name:
        return Response({"detail": "El nombre es obligatorio"}, status=400)

    m.contact_name = contact_name
    m.password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    m.status = "activo"
    m.invite_token = ""
    m.save()
    return Response(_make_user_response(m.partner, member=m), status=201)


@api_view(["PATCH", "DELETE"])
def partner_user_detail(request, user_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        m = PartnerUser.objects.get(id=user_id)
    except PartnerUser.DoesNotExist:
        return Response({"detail": "Usuario no encontrado"}, status=404)

    if user.role == "admin":
        pass
    elif user.role == "socio" and not getattr(user, "is_member", False) and m.partner_id == user.id:
        pass
    else:
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "DELETE":
        m.delete()
        return Response({"ok": True})

    role = request.data.get("role")
    status_val = request.data.get("status")
    contact_name = request.data.get("contact_name")
    if role is not None and role in ("owner", "admin", "member"):
        m.role = role
    if status_val is not None and status_val in ("invitado", "activo", "desactivado"):
        m.status = status_val
        if status_val != "activo":
            m.invite_token = ""
    if contact_name is not None:
        m.contact_name = (contact_name or "").strip()
    m.save()
    return Response(_serialize_member(m))


# ─── MDF / Co-marketing (Generate Demand) ───────────────


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _fmt_date(value):
    d = _parse_date(value)
    return d.isoformat() if d else None


def _mdf_serialize(m):
    return {
        "id": m.id,
        "partner_id": m.partner_id,
        "partner_name": m.partner.company_name,
        "title": m.title,
        "campaign_type": m.campaign_type,
        "description": m.description,
        "requested_amount": m.requested_amount,
        "approved_amount": m.approved_amount,
        "start_date": _fmt_date(m.start_date),
        "end_date": _fmt_date(m.end_date),
        "status": m.status,
        "approval_notes": m.approval_notes,
        "actual_spend": m.actual_spend,
        "report_notes": m.report_notes,
        "reported_at": m.reported_at.isoformat() if m.reported_at else None,
        "closed_at": m.closed_at.isoformat() if m.closed_at else None,
        "created_at": m.created_at.isoformat() if m.created_at else "",
    }


def _mdf_committed(partner, year=None):
    from django.db.models import Sum
    qs = MdfRequest.objects.filter(partner=partner, status__in=["aprobado", "reportado", "cerrado"])
    if year:
        qs = qs.filter(created_at__year=year)
    return qs.aggregate(total=Sum("approved_amount"))["total"] or 0.0


def _mdf_stats(user, partner_id=None):
    if user.role == "admin":
        partners = Partner.objects.exclude(role="admin").all()
        if partner_id:
            partners = partners.filter(id=partner_id)
        total_budget = sum(p.mdf_budget_year for p in partners)
        total_committed = 0.0
        by_partner = []
        for p in partners:
            committed = _mdf_committed(p)
            total_committed += committed
            by_partner.append({
                "partner_id": p.id, "partner_name": p.company_name,
                "budget": p.mdf_budget_year, "committed": committed,
                "available": max(0, p.mdf_budget_year - committed),
                "requests": MdfRequest.objects.filter(partner=p).count(),
            })
        return {
            "role": "admin",
            "total_budget": total_budget,
            "total_committed": total_committed,
            "total_available": max(0, total_budget - total_committed),
            "by_partner": by_partner,
        }
    committed = _mdf_committed(user)
    return {
        "role": "partner",
        "budget": user.mdf_budget_year,
        "committed": committed,
        "available": max(0, user.mdf_budget_year - committed),
        "requests": MdfRequest.objects.filter(partner=user).count(),
    }


@api_view(["GET", "POST"])
def mdf_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    if request.method == "GET":
        if user.role == "admin":
            qs = MdfRequest.objects.select_related("partner").all()
            pid = request.query_params.get("partner_id")
            if pid:
                qs = qs.filter(partner_id=pid)
            st = request.query_params.get("status")
            if st:
                qs = qs.filter(status=st)
        else:
            qs = MdfRequest.objects.filter(partner=user)
        return Response([_mdf_serialize(m) for m in qs])

    if user.role != "socio":
        return Response({"detail": "Solo partners pueden solicitar MDF"}, status=403)

    title = (request.data.get("title") or "").strip()
    campaign_type = request.data.get("campaign_type") or "evento"
    description = (request.data.get("description") or "").strip()
    requested_amount = float(request.data.get("requested_amount") or 0)
    start_date = _parse_date(request.data.get("start_date"))
    end_date = _parse_date(request.data.get("end_date"))
    if not title or requested_amount <= 0:
        return Response({"detail": "Título e importe obligatorios"}, status=400)

    m = MdfRequest.objects.create(
        partner=user, title=title, campaign_type=campaign_type,
        description=description, requested_amount=requested_amount,
        start_date=start_date, end_date=end_date,
    )
    return Response(_mdf_serialize(m), status=201)


@api_view(["GET"])
def mdf_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)
    pid = request.query_params.get("partner_id") if user.role == "admin" else None
    return Response(_mdf_stats(user, pid))


@api_view(["PATCH", "DELETE"])
def mdf_detail(request, mdf_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        m = MdfRequest.objects.select_related("partner").get(id=mdf_id)
    except MdfRequest.DoesNotExist:
        return Response({"detail": "Solicitud no encontrada"}, status=404)

    is_admin = user.role == "admin"
    is_owner = user.role == "socio" and m.partner_id == user.id
    if not (is_admin or is_owner):
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "DELETE":
        if is_admin or m.status in ("solicitado", "en_revision", "rechazado"):
            m.delete()
            return Response({"ok": True})
        return Response({"detail": "No se puede eliminar una solicitud en curso"}, status=400)

    action = request.data.get("action") or ""

    if is_admin:
        if action == "approve":
            if m.status not in ("solicitado", "en_revision", "rechazado"):
                return Response({"detail": "Estado no permite aprobación"}, status=400)
            amount = request.data.get("approved_amount")
            m.approved_amount = float(amount) if amount is not None else m.requested_amount
            m.approval_notes = (request.data.get("approval_notes") or "").strip()
            m.status = "aprobado"
            m.save()
            return Response(_mdf_serialize(m))
        if action == "reject":
            m.status = "rechazado"
            m.approval_notes = (request.data.get("approval_notes") or "").strip()
            m.save()
            return Response(_mdf_serialize(m))
        if action == "close":
            if m.status != "reportado":
                return Response({"detail": "Solo se cierra una solicitud reportada"}, status=400)
            m.status = "cerrado"
            m.closed_at = dj_timezone.now()
            m.save()
            return Response(_mdf_serialize(m))
        if action == "review":
            m.status = "en_revision"
            m.save()
            return Response(_mdf_serialize(m))
        return Response({"detail": "Acción no válida"}, status=400)

    # partner side
    if action == "report":
        if m.status not in ("aprobado", "reportado"):
            return Response({"detail": "Solo se reporta una solicitud aprobada"}, status=400)
        spend = request.data.get("actual_spend")
        m.actual_spend = float(spend) if spend is not None else m.approved_amount
        m.report_notes = (request.data.get("report_notes") or "").strip()
        m.reported_at = dj_timezone.now()
        m.status = "reportado"
        m.save()
        award_points(m.partner, "mdf_reportado", note=m.title, source_type="mdf", source_id=m.id)
        return Response(_mdf_serialize(m))

    if m.status not in ("solicitado", "en_revision", "rechazado"):
        return Response({"detail": "La solicitud ya no es editable"}, status=400)
    if "title" in request.data:
        m.title = (request.data.get("title") or "").strip()
    if "campaign_type" in request.data:
        m.campaign_type = request.data.get("campaign_type")
    if "description" in request.data:
        m.description = (request.data.get("description") or "").strip()
    if "requested_amount" in request.data:
        m.requested_amount = float(request.data.get("requested_amount") or 0)
    if "start_date" in request.data:
        m.start_date = _parse_date(request.data.get("start_date"))
    if "end_date" in request.data:
        m.end_date = _parse_date(request.data.get("end_date"))
    m.save()
    return Response(_mdf_serialize(m))


# ─── Conflict Management (RQ-20) ────────────────────────


def _conflict_opp_info(o):
    if not o:
        return None
    return {
        "id": o.id,
        "name": o.name,
        "company_name": o.company_name,
        "stage": o.stage,
        "amount": o.amount,
        "partner_id": o.partner_id,
        "partner_name": o.partner.company_name,
    }


def _conflict_serialize(c):
    return {
        "id": c.id,
        "company_name": c.company_name,
        "status": c.status,
        "opportunity": _conflict_opp_info(c.opportunity),
        "conflicting_opportunity": _conflict_opp_info(c.conflicting_opportunity),
        "reporter_id": c.reporter_id,
        "reporter_name": c.reporter.company_name,
        "winner_opportunity_id": c.winner_opportunity_id,
        "winner_partner_name": c.winner_opportunity.partner.company_name if c.winner_opportunity else None,
        "notes": c.notes,
        "resolution": c.resolution,
        "created_at": c.created_at.isoformat() if c.created_at else "",
        "resolved_at": c.resolved_at.isoformat() if c.resolved_at else None,
    }


def _conflict_qs_select():
    return ChannelConflict.objects.select_related(
        "opportunity__partner", "conflicting_opportunity__partner",
        "reporter", "winner_opportunity__partner",
    )


@api_view(["GET", "POST"])
def conflicts_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        if user.role == "admin":
            qs = _conflict_qs_select()
        else:
            qs = _conflict_qs_select().filter(
                Q(reporter=user) | Q(opportunity__partner=user) | Q(conflicting_opportunity__partner=user)
            )
        st = request.query_params.get("status")
        if st:
            qs = qs.filter(status=st)
        return Response([_conflict_serialize(c) for c in qs])

    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    company = (request.data.get("company_name") or "").strip()
    if not company:
        return Response({"detail": "Empresa obligatoria"}, status=400)
    notes = (request.data.get("notes") or "").strip()

    existing = ChannelConflict.objects.filter(
        company_name__iexact=company
    ).exclude(status__in=["resuelto", "cerrado"]).first()
    if existing:
        return Response(_conflict_serialize(existing), status=200)

    candidates = list(Opportunity.objects.filter(
        company_name__iexact=company
    ).exclude(stage="perdida")[:5])
    mine = [o for o in candidates if o.partner_id == user.id] if user.role == "socio" else candidates
    opp = mine[0] if mine else (candidates[0] if candidates else None)
    other = None
    if opp:
        rem = [o for o in candidates if o.id != opp.id]
        other = rem[0] if rem else None

    c = ChannelConflict.objects.create(
        company_name=company, opportunity=opp, conflicting_opportunity=other,
        reporter=user, notes=notes or "Reportado manualmente",
    )
    for o in (opp, other):
        if o and not o.conflict_indicator:
            o.conflict_indicator = True
            o.save(update_fields=["conflict_indicator"])
    notify_ids = [o.partner_id for o in (opp, other) if o and o.partner_id and o.partner_id != user.id]
    _notify(
        notify_ids, "conflicto",
        {"en": f"Channel conflict reported: {company}", "es": f"Conflicto de canal reportado: {company}", "de": f"Kanalkonflikt gemeldet: {company}"},
        {"en": "One of your opportunities may be in conflict. Review it.", "es": "Una de tus oportunidades puede estar en conflicto. Revísala.", "de": "Eine Ihrer Opportunities könnte betroffen sein. Bitte prüfen."},
        "/partner/conflicts",
    )
    return Response(_conflict_serialize(c), status=201)


@api_view(["GET", "PATCH", "DELETE"])
def conflict_detail(request, conflict_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        c = _conflict_qs_select().get(id=conflict_id)
    except ChannelConflict.DoesNotExist:
        return Response({"detail": "Conflicto no encontrado"}, status=404)

    is_admin = user.role == "admin"
    is_involved = (
        c.reporter_id == user.id
        or (c.opportunity and c.opportunity.partner_id == user.id)
        or (c.conflicting_opportunity and c.conflicting_opportunity.partner_id == user.id)
    )
    if not (is_admin or is_involved):
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "GET":
        return Response(_conflict_serialize(c))

    if request.method == "DELETE":
        if not is_admin:
            return Response({"detail": "Solo administradores"}, status=403)
        for o in (c.opportunity, c.conflicting_opportunity):
            if o and o.conflict_indicator:
                o.conflict_indicator = False
                o.save(update_fields=["conflict_indicator"])
        c.delete()
        return Response({"ok": True})

    if not is_admin:
        return Response({"detail": "Solo administradores"}, status=403)

    action = request.data.get("action") or ""
    if "notes" in request.data:
        c.notes = (request.data.get("notes") or "").strip()

    if action == "review":
        c.status = "en_resolucion"
        c.save()
        return Response(_conflict_serialize(c))

    if action == "open":
        c.status = "abierto"
        c.save()
        return Response(_conflict_serialize(c))

    if action == "resolve":
        if c.status in ("resuelto", "cerrado"):
            return Response({"detail": "Conflicto ya resuelto"}, status=400)
        winner_id = request.data.get("winner_opportunity_id")
        involved = [o for o in (c.opportunity, c.conflicting_opportunity) if o]
        winner = next((o for o in involved if o.id == winner_id), None)
        if not winner:
            return Response({"detail": "Indica la oportunidad ganadora"}, status=400)
        today = dj_timezone.localdate()
        for o in involved:
            if o.id == winner.id:
                o.conflict_indicator = False
                o.protection_end_date = today + timedelta(days=90)
                o.save(update_fields=["conflict_indicator", "protection_end_date"])
            else:
                o.conflict_indicator = False
                o.protection_end_date = None
                o.save(update_fields=["conflict_indicator", "protection_end_date"])
        c.status = "resuelto"
        c.winner_opportunity = winner
        c.resolution = (request.data.get("resolution") or "").strip() or f"Gana {winner.partner.company_name}"
        c.resolved_at = dj_timezone.now()
        c.save()
        involved_partners = [o.partner_id for o in involved if o.partner_id]
        winner_label = winner.partner.company_name
        _notify(
            involved_partners, "conflicto",
            {"en": f"Conflict resolved: {c.company_name}", "es": f"Conflicto resuelto: {c.company_name}", "de": f"Konflikt gelöst: {c.company_name}"},
            {"en": f"The winning partner is {winner_label}", "es": f"El partner ganador es {winner_label}", "de": f"Der Gewinner-Partner ist {winner_label}"},
            "/partner/conflicts",
        )
        return Response(_conflict_serialize(c))

    if action == "close":
        if c.status in ("resuelto", "cerrado"):
            return Response({"detail": "Conflicto ya cerrado"}, status=400)
        for o in (c.opportunity, c.conflicting_opportunity):
            if o and o.conflict_indicator:
                o.conflict_indicator = False
                o.save(update_fields=["conflict_indicator"])
        c.status = "cerrado"
        c.resolution = (request.data.get("resolution") or "").strip() or "Cerrado sin asignación"
        c.resolved_at = dj_timezone.now()
        c.save()
        return Response(_conflict_serialize(c))

    return Response({"detail": "Acción no válida"}, status=400)


@api_view(["GET"])
def conflicts_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if user.role == "admin":
        qs = ChannelConflict.objects.all()
        by_status = {s: qs.filter(status=s).count() for s in ["abierto", "en_resolucion", "resuelto", "cerrado"]}
        by_company = (
            qs.values("company_name").annotate(count=Count("id")).order_by("-count", "company_name")[:12]
        )
        return Response({
            "role": "admin",
            "total": qs.count(),
            "open": by_status["abierto"] + by_status["en_resolucion"],
            "by_status": by_status,
            "by_company": [{"company_name": x["company_name"], "count": x["count"]} for x in by_company],
        })

    qs = ChannelConflict.objects.filter(
        Q(reporter=user) | Q(opportunity__partner=user) | Q(conflicting_opportunity__partner=user)
    )
    return Response({
        "role": "socio",
        "total": qs.count(),
        "open": qs.filter(status__in=["abierto", "en_resolucion"]).count(),
        "resolved": qs.filter(status="resuelto").count(),
        "closed": qs.filter(status="cerrado").count(),
    })


# ─── Engage / Comunicaciones (RNV) ──────────────────────


def _communication_serialize(comm, user=None):
    d = {
        "id": comm.id,
        "subject": comm.subject,
        "body": comm.body,
        "channel": comm.channel,
        "audience": comm.audience,
        "sent_by_name": comm.sent_by.company_name if comm.sent_by else "",
        "created_at": comm.created_at.isoformat() if comm.created_at else "",
        "recipient_count": comm.recipients.count(),
        "read_count": comm.recipients.filter(read=True).count(),
    }
    if user and user.role == "socio":
        rec = comm.recipients.filter(partner=user).first()
        d["read"] = rec.read if rec else None
        d["read_at"] = rec.read_at.isoformat() if rec and rec.read_at else None
    return d


@api_view(["GET", "POST"])
def communications_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        if user.role == "admin":
            qs = Communication.objects.all()
            return Response([_communication_serialize(c) for c in qs])
        recs = CommunicationRecipient.objects.filter(partner=user)
        return Response([_communication_serialize(r.communication, user) for r in recs])

    if user.role != "admin":
        return Response({"detail": "Solo administradores"}, status=403)
    subject = (request.data.get("subject") or "").strip()
    body = (request.data.get("body") or "").strip()
    if not subject or not body:
        return Response({"detail": "Subject y body son obligatorios"}, status=400)
    channel = request.data.get("channel") or "in_app"
    if channel not in ["in_app", "email"]:
        return Response({"detail": "Canal no válido"}, status=400)
    audience = request.data.get("audience") or "todos"
    if audience not in ["todos", "seleccion"]:
        return Response({"detail": "Audiencia no válida"}, status=400)
    partner_ids = request.data.get("partner_ids") or []
    if audience == "seleccion" and not partner_ids:
        return Response({"detail": "Selecciona al menos un partner"}, status=400)

    comm = Communication.objects.create(subject=subject, body=body, channel=channel, audience=audience, sent_by=user)
    if audience == "todos":
        partners = Partner.objects.filter(status="activo")
    else:
        partners = Partner.objects.filter(id__in=partner_ids)
    for p in partners:
        CommunicationRecipient.objects.get_or_create(communication=comm, partner=p)
    _notify(
        [p.id for p in partners], "comunicacion",
        {"en": subject, "es": subject, "de": subject},
        {"en": body[:160], "es": body[:160], "de": body[:160]},
        "/partner",
    )
    return Response(_communication_serialize(comm, user), status=201)


@api_view(["GET", "PATCH"])
def communication_detail(request, comm_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    try:
        comm = Communication.objects.get(id=comm_id)
    except Communication.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)
    is_admin = user.role == "admin"

    if not is_admin and not comm.recipients.filter(partner=user).exists():
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "GET":
        if is_admin:
            recipients = [
                {"partner_id": r.partner_id, "partner_name": r.partner.company_name,
                 "read": r.read, "read_at": r.read_at.isoformat() if r.read_at else None}
                for r in comm.recipients.all()
            ]
            return Response(dict(_communication_serialize(comm, user), recipients=recipients))
        return Response(_communication_serialize(comm, user))

    # PATCH: partner marks as read
    if is_admin:
        return Response({"detail": "Acción no válida"}, status=400)
    rec = comm.recipients.filter(partner=user).first()
    if request.data.get("action") == "read" and not rec.read:
        rec.read = True
        rec.read_at = dj_timezone.now()
        rec.save(update_fields=["read", "read_at"])
    return Response(_communication_serialize(comm, user))


@api_view(["GET"])
def communications_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if user.role == "admin":
        qs = Communication.objects.all()
        recs = CommunicationRecipient.objects.filter(communication__in=qs)
        delivered = recs.count()
        read = recs.filter(read=True).count()
        recent = [
            {
                "id": c.id,
                "subject": c.subject,
                "channel": c.channel,
                "created_at": c.created_at.isoformat() if c.created_at else "",
                "recipient_count": c.recipients.count(),
                "read_count": c.recipients.filter(read=True).count(),
            }
            for c in qs[:10]
        ]
        return Response({
            "role": "admin",
            "total": qs.count(),
            "delivered": delivered,
            "read": read,
            "read_rate": round(read / delivered * 100) if delivered else 0,
            "recent": recent,
        })

    recs = CommunicationRecipient.objects.filter(partner=user)
    return Response({
        "role": "socio",
        "total": recs.count(),
        "unread": recs.filter(read=False).count(),
    })


# ─── Rewards / Motivate (RNV) ───────────────────────────


def _reward_serialize(r, user=None):
    d = {
        "id": r.id,
        "name": r.name,
        "description": r.description,
        "category": r.category,
        "points_cost": r.points_cost,
        "stock": r.stock,
        "active": r.active,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    }
    if user and user.role == "socio":
        d["can_afford"] = user.points_balance >= r.points_cost and (r.stock is None or r.stock > 0)
    return d


def _redemption_serialize(red, with_partner=False):
    d = {
        "id": red.id,
        "partner_id": red.partner_id,
        "reward_id": red.reward_id,
        "reward_name": red.reward.name,
        "reward_category": red.reward.category,
        "points_spent": red.points_spent,
        "status": red.status,
        "admin_notes": red.admin_notes,
        "created_at": red.created_at.isoformat() if red.created_at else "",
    }
    if with_partner:
        d["partner_name"] = red.partner.company_name
    return d


@api_view(["GET", "POST"])
def rewards_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        if user.role == "admin":
            qs = Reward.objects.all()
        else:
            if user.status != "activo":
                return Response({"detail": "Solo partners aprobados"}, status=403)
            qs = Reward.objects.filter(active=True)
        return Response([_reward_serialize(r, user) for r in qs])

    if user.role != "admin":
        return Response({"detail": "Solo administradores"}, status=403)

    name = request.data.get("name")
    if not name or not (name.get("en") or "").strip():
        return Response({"detail": "Nombre obligatorio"}, status=400)
    points_cost = int(request.data.get("points_cost") or 0)
    if points_cost <= 0:
        return Response({"detail": "El coste en puntos debe ser mayor que 0"}, status=400)
    stock_raw = request.data.get("stock")
    stock = None
    if stock_raw not in (None, ""):
        stock = max(0, int(stock_raw))
    r = Reward.objects.create(
        name=name,
        description=request.data.get("description") or {},
        category=request.data.get("category") or "merchandising",
        points_cost=points_cost,
        stock=stock,
    )
    return Response(_reward_serialize(r, user), status=201)


@api_view(["PATCH", "DELETE"])
def reward_detail(request, reward_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role != "admin":
        return Response({"detail": "Solo administradores"}, status=403)

    try:
        r = Reward.objects.get(id=reward_id)
    except Reward.DoesNotExist:
        return Response({"detail": "Recompensa no encontrada"}, status=404)

    if request.method == "DELETE":
        r.active = False
        r.save(update_fields=["active"])
        return Response({"ok": True})

    if "name" in request.data and request.data.get("name"):
        r.name = request.data.get("name")
    if "description" in request.data:
        r.description = request.data.get("description") or {}
    if "category" in request.data:
        r.category = request.data.get("category")
    if "points_cost" in request.data:
        pc = int(request.data.get("points_cost") or 0)
        if pc <= 0:
            return Response({"detail": "El coste en puntos debe ser mayor que 0"}, status=400)
        r.points_cost = pc
    if "stock" in request.data:
        stock_raw = request.data.get("stock")
        r.stock = None if stock_raw in (None, "") else max(0, int(stock_raw))
    if "active" in request.data:
        r.active = bool(request.data.get("active"))
    r.save()
    return Response(_reward_serialize(r, user))


@api_view(["GET", "POST"])
def redemptions_list_or_create(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if request.method == "GET":
        if user.role == "admin":
            qs = RewardRedemption.objects.select_related("partner", "reward")
        else:
            qs = RewardRedemption.objects.filter(partner=user).select_related("reward")
        return Response([_redemption_serialize(red, with_partner=user.role == "admin") for red in qs])

    if user.role == "socio" and user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)

    reward_id = request.data.get("reward_id")
    try:
        r = Reward.objects.get(id=reward_id)
    except Reward.DoesNotExist:
        return Response({"detail": "Recompensa no encontrada"}, status=404)
    if not r.active:
        return Response({"detail": "Recompensa no disponible"}, status=400)
    if r.stock is not None and r.stock <= 0:
        return Response({"detail": "Recompensa agotada"}, status=400)

    name = r.name.get("en") or r.name.get("es") or r.name.get("de") or "Reward"
    if not spend_points(user, r.points_cost, note=name, source_type="reward", source_id=r.id):
        return Response({"detail": "Puntos insuficientes"}, status=400)

    red = RewardRedemption.objects.create(partner=user, reward=r, points_spent=r.points_cost)
    if r.stock is not None:
        r.stock -= 1
        r.save(update_fields=["stock"])
    return Response(_redemption_serialize(red), status=201)


@api_view(["PATCH", "DELETE"])
def redemption_detail(request, redemption_id):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    try:
        red = RewardRedemption.objects.select_related("partner", "reward").get(id=redemption_id)
    except RewardRedemption.DoesNotExist:
        return Response({"detail": "Canje no encontrado"}, status=404)

    is_admin = user.role == "admin"
    is_owner = user.role == "socio" and red.partner_id == user.id
    if not (is_admin or is_owner):
        return Response({"detail": "Access denied"}, status=403)

    if request.method == "DELETE":
        if red.status != "solicitado":
            return Response({"detail": "Solo se puede cancelar un canje solicitado"}, status=400)
        refund_points(red.partner, red.points_spent, note=f"Cancel: {red.reward.name.get('en') or 'reward'}")
        red.delete()
        return Response({"ok": True})

    if not is_admin:
        return Response({"detail": "Solo administradores"}, status=403)

    action = request.data.get("action") or ""
    if action == "deliver":
        if red.status != "solicitado":
            return Response({"detail": "Estado no permite entrega"}, status=400)
        red.status = "entregado"
    elif action == "reject":
        if red.status != "solicitado":
            return Response({"detail": "Estado no permite rechazo"}, status=400)
        red.status = "rechazado"
        refund_points(red.partner, red.points_spent, note=f"Rejected: {red.reward.name.get('en') or 'reward'}")
    else:
        return Response({"detail": "Acción no válida"}, status=400)
    red.admin_notes = (request.data.get("admin_notes") or "").strip()
    red.save()
    return Response(_redemption_serialize(red, with_partner=True))


@api_view(["GET"])
def rewards_stats(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    if user.role == "admin":
        from django.db.models import Sum as SumF
        tops = (
            Partner.objects.annotate(total_earned=SumF("points_earned"))
            .order_by("-points_balance")[:10]
        )
        return Response({
            "role": "admin",
            "top_partners": [
                {"partner_id": p.id, "partner_name": p.company_name,
                 "balance": p.points_balance, "earned": p.points_earned}
                for p in tops
            ],
            "pending_redemptions": RewardRedemption.objects.filter(status="solicitado").count(),
            "total_delivered": RewardRedemption.objects.filter(status="entregado").count(),
            "active_rewards": Reward.objects.filter(active=True).count(),
        })

    if user.status != "activo":
        return Response({"detail": "Solo partners aprobados"}, status=403)
    tx = PointTransaction.objects.filter(partner=user)
    spent = abs(sum(t.amount for t in tx if t.amount < 0))
    return Response({
        "role": "socio",
        "balance": user.points_balance,
        "earned": user.points_earned,
        "spent": spent,
        "redemptions": RewardRedemption.objects.filter(partner=user).count(),
        "pending_redemptions": RewardRedemption.objects.filter(partner=user, status="solicitado").count(),
        "recent": [point_serialize(t) for t in tx[:10]],
    })


@api_view(["POST"])
def rewards_adjust_points(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)
    if user.role != "admin":
        return Response({"detail": "Solo administradores"}, status=403)

    partner_id = request.data.get("partner_id")
    partner = Partner.objects.filter(id=partner_id).first()
    if not partner:
        return Response({"detail": "Partner no encontrado"}, status=404)
    note = (request.data.get("note") or "").strip()
    try:
        amount = int(request.data.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    tx = adjust_points(partner, amount, note=note)
    if not tx:
        return Response({"detail": "Importe no válido o saldo insuficiente"}, status=400)
    return Response({
        "transaction": point_serialize(tx),
        "balance": partner.points_balance,
    })


# ─── Helper ─────────────────────────────────────────────

def _current_user(request):
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if not auth_header.startswith("Bearer "):
        return None
    token = auth_header.split(" ", 1)[1]
    try:
        from rest_framework_simplejwt.tokens import AccessToken
        at = AccessToken(token)
        if TokenBlacklist.objects.filter(jti=at["jti"]).exists():
            return None
        user_id = at.get("user_id")
        partner = Partner.objects.get(id=user_id)
        member_id = at.get("member_id")
        if member_id:
            member = PartnerUser.objects.filter(id=member_id, status="activo").first()
            if not member or member.partner_id != partner.id:
                return None
            partner.is_member = True
            partner.member = member
        else:
            partner.is_member = False
            partner.member = None
        return partner
    except Exception:
        return None


# ─── Cost exporter (ARR calculator -> Excel) ───────────

_COST_DEFAULT_TEXTS = {
    "title": {"en": "Cost Indication", "es": "Indicación de costos", "de": "Kostenindikation"},
    "subtitle": {
        "en": "Estimated annual costs for the proposed products",
        "es": "Costes anuales estimados para los productos propuestos",
        "de": "Geschätzte jährliche Kosten für die vorgeschlagenen Produkte",
    },
    "footer": {
        "en": "This is an indicative, non-binding cost estimate prepared by the partner.",
        "es": "Esta es una estimación de costos orientativa y no vinculante preparada por el partner.",
        "de": "Dies ist eine unverbindliche, indikative Kostenschätzung des Partners.",
    },
    "product_col": {"en": "Product", "es": "Producto", "de": "Produkt"},
    "annual_col": {"en": "Annual license cost", "es": "Coste anual de licencia", "de": "Jährliche Lizenzkosten"},
}


def _get_cost_settings():
    obj = CostExportSetting.objects.first()
    if not obj:
        obj = CostExportSetting.objects.create(
            title=_COST_DEFAULT_TEXTS["title"],
            subtitle=_COST_DEFAULT_TEXTS["subtitle"],
            footer=_COST_DEFAULT_TEXTS["footer"],
            product_col=_COST_DEFAULT_TEXTS["product_col"],
            annual_col=_COST_DEFAULT_TEXTS["annual_col"],
        )
    return obj


def _cost_settings_payload(obj):
    payload = {}
    for f in ("title", "subtitle", "footer", "product_col", "annual_col"):
        payload[f] = getattr(obj, f) if isinstance(getattr(obj, f), dict) else {}
        if not payload[f]:
            payload[f] = _COST_DEFAULT_TEXTS[f]
    return payload


def _text_multi(raw, lang, key):
    d = raw if isinstance(raw, dict) else {}
    for k in (lang, "es", "en"):
        if d.get(k):
            return d[k]
    return key


@api_view(["GET", "PUT"])
def calculator_settings(request):
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    obj = _get_cost_settings()
    if request.method == "PUT":
        if user.role != "admin":
            return Response({"detail": "Admin access required"}, status=403)
        data = request.data or {}
        for f in ("title", "subtitle", "footer", "product_col", "annual_col"):
            if f in data:
                setattr(obj, f, data[f])
        obj.save()
        return Response(_cost_settings_payload(obj))

    return Response(_cost_settings_payload(obj))


@api_view(["POST"])
def calculator_export(request):
    """
    Generate an .xlsx 'cost indication' from the ARR calculator.
    Body: { products: [{key, name, price, annual, m12, m24, m36, currency}],
            company_name, company_size, currency, total12, total24, total36 }
    Applies admin-configured fixed texts.
    """
    user = _current_user(request)
    if not user:
        return Response({"detail": "Unauthorized"}, status=401)

    import io
    from openpyxl import Workbook

    lang = request.query_params.get("lang", "es")
    data = request.data or {}
    settings = _get_cost_settings()
    symbol_map = {"eur": "€", "chf": "CHF", "usd": "$", "otro": "$"}
    fmt_money = lambda n, cur: f"{symbol_map.get(cur, '$')} {round((float(n) or 0), 2):,.2f}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Costos"

    title = _text_multi(settings.title, lang, "Cost Indication")
    subtitle = _text_multi(settings.subtitle, lang, "")
    footer = _text_multi(settings.footer, lang, "")
    product_col = _text_multi(settings.product_col, lang, "Product")
    annual_col = _text_multi(settings.annual_col, lang, "Annual license cost")

    company = (data.get("company_name") or "").strip() or user.company_name or ""
    company_size = data.get("company_size") or ""
    currency = data.get("currency") or currency_for_country(user.country)

    ws["A1"] = title
    ws["A1"].font = openpyxl_font(bold=True, size=14)
    ws["A2"] = subtitle
    ws["A2"].font = openpyxl_font(italic=True, color="666666")
    r = 4
    if company:
        ws.cell(row=r, column=1, value=f"{company}  ({company_size})").font = openpyxl_font(bold=True)
        r += 1
    r += 1

    head = r
    ws.cell(row=head, column=1, value=product_col).font = openpyxl_font(bold=True)
    ws.cell(row=head, column=2, value=annual_col).font = openpyxl_font(bold=True)
    ws.cell(row=head, column=3, value="12m").font = openpyxl_font(bold=True)
    ws.cell(row=head, column=4, value="24m").font = openpyxl_font(bold=True)
    ws.cell(row=head, column=5, value="36m").font = openpyxl_font(bold=True)
    for c in range(1, 6):
        ws.cell(row=head, column=c).fill = openpyxl_fill("DDEEF7")

    rr = head + 1
    total12 = 0.0
    total24 = 0.0
    total36 = 0.0
    for p in data.get("products", []):
        row_cur = p.get("currency") or currency
        ws.cell(row=rr, column=1, value=p.get("name") or p.get("key") or "")
        ws.cell(row=rr, column=2, value=round((float(p.get("price") or 0)), 2))
        ws.cell(row=rr, column=2).number_format = "#,##0.00"
        ws.cell(row=rr, column=3, value=round((float(p.get("m12") or 0)), 2)).number_format = "#,##0.00"
        ws.cell(row=rr, column=4, value=round((float(p.get("m24") or 0)), 2)).number_format = "#,##0.00"
        ws.cell(row=rr, column=5, value=round((float(p.get("m36") or 0)), 2)).number_format = "#,##0.00"
        total12 += float(p.get("m12") or 0)
        total24 += float(p.get("m24") or 0)
        total36 += float(p.get("m36") or 0)
        rr += 1

    if data.get("total12") is not None:
        total12 = float(data["total12"] or 0)
    if data.get("total24") is not None:
        total24 = float(data["total24"] or 0)
    if data.get("total36") is not None:
        total36 = float(data["total36"] or 0)

    ws.cell(row=rr, column=1, value="TOTAL").font = openpyxl_font(bold=True)
    ws.cell(row=rr, column=3, value=round(total12, 2)).font = openpyxl_font(bold=True)
    ws.cell(row=rr, column=3).number_format = "#,##0.00"
    ws.cell(row=rr, column=4, value=round(total24, 2)).font = openpyxl_font(bold=True)
    ws.cell(row=rr, column=4).number_format = "#,##0.00"
    ws.cell(row=rr, column=5, value=round(total36, 2)).font = openpyxl_font(bold=True)
    ws.cell(row=rr, column=5).number_format = "#,##0.00"
    ws.cell(row=rr, column=2, value=round(sum(float(p.get("price") or 0) for p in data.get("products", [])), 2)).font = openpyxl_font(bold=True)
    ws.cell(row=rr, column=2).number_format = "#,##0.00"

    if footer:
        frow = rr + 2
        ws.cell(row=frow, column=1, value=footer).font = openpyxl_font(italic=True, color="888888")

    for col, width in zip("ABCDE", [30, 22, 14, 14, 14]):
        ws.column_dimensions[col].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    from django.http import HttpResponse
    resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="cost_indication.xlsx"'
    return resp


def openpyxl_font(bold=False, italic=False, size=11, color="000000"):
    from openpyxl.styles import Font
    return Font(bold=bold, italic=italic, size=size, color=color)


def openpyxl_fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
